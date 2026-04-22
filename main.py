import os
import sys
import json
import time
from datetime import datetime

from openpyxl import Workbook, load_workbook
from google import genai
from google.genai import types  # for audio parts

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, "Scenario EIMY.xlsx")
SENTIMENT_SHEET_NAME = "Sentiment_Results"
SCENARIO_SHEET_NAME = "Scenario"


# Folder where we will store .txt transcripts
TRANSCRIPT_DIR = os.path.join(BASE_DIR, "transcripts")
os.makedirs(TRANSCRIPT_DIR, exist_ok=True)

# Will hold scenarios loaded from Excel
SCENARIOS = []

# === CONFIG FLAGS (change these to save quota) ===
ENABLE_FORMALISATION = True       # True = always clean transcript
ENABLE_TRANSLATION = False        # False = skip English, True = translate

# Get API key from environment variable
API_KEY = os.getenv("GEMINI_API_KEY")
if not API_KEY:
    raise RuntimeError("GEMINI_API_KEY environment variable is not set.")

# Create Gemini client
client = genai.Client(api_key=API_KEY)
MODEL_NAME = "gemini-2.5-flash"

SYSTEM_PROMPT = """
You are a sentiment and complaint detection engine for client feedback.

Task:
- Read the client's feedback text (already transcribed if it came from audio).
- Optionally, you may also be given the raw audio of the client speaking.
- Decide if the feedback is:
  - "Complaint" -> The client is unhappy, reporting problems, expressing frustration or dissatisfaction.
  - "Non-complaint" -> The feedback is generally positive or just informative with no clear problem.
  - "Neutral" -> Mixed or unclear sentiment, not clearly complaint or positive.
- If audio is provided, use BOTH:
  - The actual words spoken
  - The tone of voice (angry, calm, annoyed, frustrated, sad, polite, stressed, etc.)
- Detect the overall tone of voice and summarise it with ONE or TWO words only.
- Give a short explanation (1–2 sentences).
- Return JSON ONLY in this exact format:

{
  "sentiment": "...",
  "score": 0,
  "tone": "...",
  "explanation": "..."
}

Where:
- "sentiment" must be exactly one of: "Complaint", "Non-complaint", "Neutral"
- "score" must be an INTEGER between 0 and 100.
- NEVER return decimal values. NEVER return floating point numbers.
- Always return whole numbers only (e.g., 0, 36, 72, 100).
- "tone" is a short description of the client's tone of voice, e.g. "angry", "calm", "frustrated", "polite", "mixed", etc.
"""

TRANSCRIPTION_PROMPT = """
You are a transcription engine for client service calls.

In each call there are two speakers:
- the client (the caller)
- the customer service staff (the agent)

The speaker may mix:
- Mandarin Chinese
- Hokkien (Minnan) dialect / Penang Hokkien slang
- English
- Malay

Transcription rules:
- Transcribe Mandarin in Chinese characters.
- Transcribe Hokkien using common informal romanization/slang
  (e.g. "sibeh sian", "bo ho liao", "tulan").
- Do NOT translate to another language.
- Preserve code-mixing exactly as spoken (Mandarin + Hokkien + English + Malay).
- Identify who is speaking from the conversation context.
- Format the output as a dialogue with speaker labels, for example:

  Client: ...
  CS: ...

- Put each utterance on its own line.
- Use exactly these labels: "Client:" for the caller, "CS:" for the customer service staff.
- Do NOT add timestamps, brackets, or explanations.
- Do NOT add extra commentary before or after the dialogue.

Return ONLY the dialogue lines.
"""

TRANSLATION_PROMPT = """
Translate the following text into natural English.

The text is a dialogue between "Client:" and "CS:" and may mix:
- Mandarin (Chinese characters)
- Hokkien/Minnan slang (romanized)
- English
- Malay

Rules:
- Preserve the speaker labels "Client:" and "CS:".
- Preserve line breaks so the conversation stays easy to read.
- Maintain the original meaning as much as possible.
- Do NOT add explanations, notes, or commentary.
- Return ONLY the translated dialogue in English with the same labels.
"""

FORMALISE_PROMPT = """
You are a linguistic cleanup engine.

Rewrite the following transcript into a clean, formal, grammatically correct version.

Rules:
- The input is a dialogue between "Client:" and "CS:".
- KEEP the speaker labels ("Client:" and "CS:") and the order of the turns.
- Do NOT merge two speakers' lines into one.
- Remove filler words such as "lah", "lor", "leh", "meh", "mah", "loh", etc.
- Remove unnecessary slang or vulgar Hokkien expressions unless they convey real meaning.
- Keep the meaning EXACTLY the same.
- If the text is Mandarin, keep it in Mandarin but with formal sentence structure.
- If the text mixes languages, you may standardise into one main language while preserving meaning.
- Do NOT add explanations, comments or extra sentences.
- Return ONLY the cleaned, formalised dialogue with the same "Client:" / "CS:" labels.
"""


# ========================
# Helper: normalise score
# ========================
def normalize_score(raw_score) -> int:
    """
    Normalise the model's 'score' field into an integer percentage 0–100.

    Handles:
    - float in [0, 1]  -> multiplied by 100 (0.9 -> 90)
    - int              -> returned as-is (clamped 0–100)
    - numeric strings  -> parsed then treated like above
    Falls back to 0 if parsing fails.
    """
    try:
        # If it's not a number yet, try to parse from string
        if not isinstance(raw_score, (int, float)):
            raw_score = float(raw_score)

        # If it's a float between 0 and 1, treat as probability
        if isinstance(raw_score, float) and 0.0 <= raw_score <= 1.0:
            value = int(round(raw_score * 100))
        else:
            value = int(round(raw_score))

        # Clamp to [0, 100] just in case
        if value < 0:
            value = 0
        if value > 100:
            value = 100

        return value
    except Exception:
        return 0


# ========================
# Gemini wrapper
# ========================
def safe_generate_content(*args, **kwargs):
    """
    Wrapper to call Gemini and automatically handle 503 (model overloaded).
    Retries 5 times with increasing wait delays.
    """
    max_retries = 5
    delay = 3  # seconds

    for attempt in range(max_retries):
        try:
            return client.models.generate_content(*args, **kwargs)

        except Exception as e:
            error_text = str(e)

            # If overloaded / unavailable
            if "UNAVAILABLE" in error_text or "503" in error_text:
                print(f"[Retry {attempt+1}/{max_retries}] Gemini is overloaded. Waiting {delay} seconds...")
                time.sleep(delay)
                delay += 2  # increase wait time
                continue

            # If quota issue handled earlier
            if "RESOURCE_EXHAUSTED" in error_text or "quota" in error_text.lower():
                print("[Quota Error] Quota exceeded.")
                raise

            # Other errors
            raise

    # If still failing after retries:
    raise RuntimeError("Gemini server still unavailable after multiple retries.")


# ========================
# Core model calls
# ========================
def analyze_sentiment(text: str) -> dict:
    """
    Send text to Gemini and return the JSON response (text-only mode).
    This will still include a 'tone' field, but inferred only from wording.
    """
    prompt = SYSTEM_PROMPT + f'\n\nClient feedback text:\n"""{text}"""'

    response = safe_generate_content(
        model=MODEL_NAME,
        contents=prompt,
        config={"response_mime_type": "application/json"},
    )

    try:
        return json.loads(response.text)
    except json.JSONDecodeError:
        print("[Error] Model did not return valid JSON (text mode):")
        print(response.text)
        raise


def translate_to_english(text: str) -> str:
    """
    Translate mixed Mandarin/Hokkien/English/Malay transcript into English.
    """
    prompt = TRANSLATION_PROMPT + f'\n\nText:\n"""{text}"""'

    response = safe_generate_content(
        model=MODEL_NAME,
        contents=prompt,
    )

    return response.text.strip()


def formalise_transcript(text: str) -> str:
    """
    Convert raw mixed-language transcript into a clean formal version.
    Removes fillers and unnecessary slang while preserving meaning.
    """
    prompt = FORMALISE_PROMPT + f'\n"""{text}"""'

    response = safe_generate_content(
        model=MODEL_NAME,
        contents=prompt,
    )

    return response.text.strip()


def transcribe_audio_to_text(audio_path: str) -> dict:
    """
    Use Gemini to:
      1) Transcribe a .wav audio file to raw text        (always)
      2) (optional) Formalise the transcript             (if ENABLE_FORMALISATION)
      3) (optional) Translate the chosen version         (if ENABLE_TRANSLATION)

    We expose ONLY one transcript to the rest of the code:
      - "transcript" = formal version if available, otherwise raw

    Returns:
        {
          "transcript": <formal-or-raw transcript>,
          "formal": <formal transcript or None>,
          "english": <English translation or None>
        }
    """
    if not os.path.exists(audio_path):
        raise FileNotFoundError(f"Audio file not found: {audio_path}")

    # Read audio bytes
    with open(audio_path, "rb") as f:
        audio_bytes = f.read()

    audio_part = types.Part.from_bytes(
        data=audio_bytes,
        mime_type="audio/wav"
    )

    print(f"\n[Info] Sending audio to Gemini for multilingual transcription: {audio_path}")

    # 1) Transcription (always happens)
    response = safe_generate_content(
        model=MODEL_NAME,
        contents=[
            TRANSCRIPTION_PROMPT,
            audio_part,
        ],
    )

    raw_transcript = (response.text or "").strip()

    # 2) Optional formalisation
    formal = None
    if ENABLE_FORMALISATION and raw_transcript:
        try:
            print("[Info] Formalising transcript...")
            formal = formalise_transcript(raw_transcript).strip()
        except Exception as e:
            formal = None
            print("[Warning] Formalisation failed, using raw transcript:", e)

    # Choose which text to expose as "transcript":
    canonical_transcript = (formal or raw_transcript or "").strip()

    # 3) Optional translation (uses the canonical transcript: usually the formal one)
    english = None
    if ENABLE_TRANSLATION and canonical_transcript:
        try:
            print("[Info] Translating transcript into English...")
            english = translate_to_english(canonical_transcript)
        except Exception as e:
            english = None
            print("[Warning] English translation failed:", e)

    return {
        "transcript": canonical_transcript,
        "formal": formal,
        "english": english
    }


def analyze_sentiment_from_audio(audio_path: str, transcript: str | None = None) -> dict:
    """
    Analyze sentiment using BOTH:
    - the raw audio (tone of voice)
    - and the transcript text (if provided)
    """
    if not os.path.exists(audio_path):
        raise FileNotFoundError(f"Audio file not found: {audio_path}")

    with open(audio_path, "rb") as f:
        audio_bytes = f.read()

    audio_part = types.Part.from_bytes(
        data=audio_bytes,
        mime_type="audio/wav"
    )

    prompt = SYSTEM_PROMPT + """

Now you will receive the client's audio recording.
The client may speak a mix of Mandarin Chinese, Hokkien (Minnan) dialect, English and Malay
(code-switching in the same sentence is possible).

If a transcript is provided below, use it as reference for the exact words.
However, you MUST also listen to the tone of voice (angry, calm, annoyed, polite, stressed, etc.)
when deciding if this is a Complaint, Non-complaint or Neutral, and when setting the "tone" field.

Do NOT include the transcript in your output.
Return ONLY the JSON as specified.
"""

    if transcript:
        prompt += f"""

Here is the transcript of the audio for reference (DO NOT reprint it):
\"\"\"{transcript}\"\"\""""

    response = safe_generate_content(
        model=MODEL_NAME,
        contents=[
            prompt,
            audio_part,
        ],
        config={"response_mime_type": "application/json"},
    )

    try:
        return json.loads(response.text)
    except json.JSONDecodeError:
        print("[Error] Model did not return valid JSON (audio mode):")
        print(response.text)
        raise


# ========================
# Excel saving
# ========================
def save_result_to_excel(result: dict, audio_path: str, transcript: str, scenario_info: dict):
    sentiment = result.get("sentiment", "")
    raw_score = result.get("score", 0)
    score = normalize_score(raw_score)
    tone = result.get("tone", "")
    explanation = result.get("explanation", "")

    scenario_id = scenario_info.get("scenario_id")
    scenario_title = scenario_info.get("scenario_title", "Unknown")

    try:
        wb = load_workbook(EXCEL_PATH)
    except FileNotFoundError:
        print(f"[Excel] Workbook '{EXCEL_PATH}' not found. Skipping Excel export.")
        return

    # ✅ Get or create sheet
    if SENTIMENT_SHEET_NAME in wb.sheetnames:
        ws = wb[SENTIMENT_SHEET_NAME]
    else:
        ws = wb.create_sheet(SENTIMENT_SHEET_NAME)

    # ✅ Ensure header row exists
    if ws.max_row == 1 and ws["A1"].value is None:
        ws.append([
            "Date & Time",
            "Audio File",
            "Sentiment",
            "Score (%)",
            "Tone",
            "Explanation",
            "Scenario ID",
            "Scenario Title",
            "Transcript",
            "Comment",
        ])

    # ✅ Ensure COMMENT header exists even for older Excel files
    if ws.cell(row=1, column=10).value is None:
        ws.cell(row=1, column=10).value = "Comment"

    now = datetime.now()
    audio_file_name = os.path.basename(audio_path)

    ws.append([
        now,
        audio_file_name,
        sentiment,
        score,
        tone,
        explanation,
        scenario_id,
        scenario_title,
        transcript,
        "",  # default empty comment
    ])

    try:
        wb.save(EXCEL_PATH)
        print("[Excel] Saved result into Excel.")
    except PermissionError:
        print(f"[Excel] Permission denied when saving to '{EXCEL_PATH}'. Close Excel and try again.")

# ========================
# Scenario loading & classification
# ========================
def load_scenarios_from_excel():
    """
    Load scenario categories from the 'Scenario' sheet in EXCEL_PATH
    into a global SCENARIOS list.

    Each scenario is a dict:
      { "id": int, "title": str, "description": str }
    """
    global SCENARIOS

    if SCENARIOS:  # already loaded
        return

    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)
    except FileNotFoundError:
        print(f"[Scenario] Workbook '{EXCEL_PATH}' not found. Scenario classification disabled.")
        SCENARIOS = []
        return

    if SCENARIO_SHEET_NAME not in wb.sheetnames:
        print(f"[Scenario] Sheet '{SCENARIO_SHEET_NAME}' not found. Scenario classification disabled.")
        SCENARIOS = []
        return

    ws = wb[SCENARIO_SHEET_NAME]

    scenarios = []

    # Scenario 1 is encoded in the header row as your file is structured
    s1_id = ws.cell(row=1, column=1).value
    s1_title = ws.cell(row=1, column=2).value
    s1_desc = ws.cell(row=1, column=3).value

    if s1_id is not None and s1_title:
        try:
            s1_id = int(s1_id)
        except Exception:
            pass
        scenarios.append({
            "id": s1_id,
            "title": str(s1_title).strip(),
            "description": str(s1_desc or "").strip(),
        })

    # Scenario 2..N are in the data rows (row >= 2)
    for row in range(2, ws.max_row + 1):
        sid = ws.cell(row=row, column=1).value
        title = ws.cell(row=row, column=2).value
        desc = ws.cell(row=row, column=3).value

        if sid is None or title is None:
            continue

        try:
            sid_int = int(sid)
        except Exception:
            sid_int = sid

        scenarios.append({
            "id": sid_int,
            "title": str(title).strip(),
            "description": str(desc or "").strip(),
        })

    SCENARIOS = scenarios
    print(f"[Scenario] Loaded {len(SCENARIOS)} scenario categories from Excel.")


def classify_scenario_from_transcript(transcript: str) -> dict:
    """
    Use Gemini to classify the call transcript into ONE of the scenario
    categories loaded from the Scenario sheet.

    Returns:
        {
          "scenario_id": <id or None>,
          "scenario_title": <str or "Unknown">
        }
    """
    load_scenarios_from_excel()

    if not SCENARIOS:
        return {"scenario_id": None, "scenario_title": "Unknown"}

    # Build a compact list of scenarios for the prompt
    scenario_text_parts = []
    for s in SCENARIOS:
        scenario_text_parts.append(
            f"ID {s['id']}: {s['title']}\nDescription: {s['description']}\n"
        )
    scenario_text = "\n".join(scenario_text_parts)

    prompt = f"""
You are a call categorization engine.

You will receive:
- A list of scenarios (each with ID, title, and description)
- A client service call transcript between "Client:" and "CS:"

Task:
- Choose the SINGLE best matching scenario ID for this call.
- If more than one seems possible, pick the one that MOST closely matches
  the client's main issue.
- If nothing fits at all, choose the closest match.

Return ONLY JSON in this exact format:

{{
  "scenario_id": <number>,
  "scenario_title": "..."
}}

Here are the scenarios:

{scenario_text}

Now here is the call transcript:

\"\"\"{transcript}\"\"\""""

    response = safe_generate_content(
        model=MODEL_NAME,
        contents=prompt,
        config={"response_mime_type": "application/json"},
    )

    try:
        data = json.loads(response.text)
    except json.JSONDecodeError:
        print("[Scenario] Model did not return valid JSON for scenario classification:")
        print(response.text)
        return {"scenario_id": None, "scenario_title": "Unknown"}

    scenario_id = data.get("scenario_id")
    scenario_title = data.get("scenario_title")

    # If scenario_title missing, try to map from our Excel list
    if scenario_title in (None, "", "Unknown") and scenario_id is not None:
        for s in SCENARIOS:
            if str(s["id"]) == str(scenario_id):
                scenario_title = s["title"]
                break

    return {
        "scenario_id": scenario_id,
        "scenario_title": scenario_title or "Unknown",
    }


# ========================
# UI helper for Flask
# ========================
def analyze_single_audio_for_ui(audio_path: str) -> dict:
    """
    Helper for the HTML UI.

    It:
      - transcribes the audio
      - runs sentiment using audio + transcript
      - classifies scenario
      - saves to Excel
      - RETURNS all important values in a dict for the web UI
    """
    try:
        # 1) Transcribe audio
        trans_result = transcribe_audio_to_text(audio_path)
        transcript = (trans_result.get("transcript") or "").strip()
        english = (trans_result.get("english") or "").strip() if trans_result.get("english") else ""

        if not transcript:
            return {
                "success": False,
                "error": "Transcript is empty. Please check the audio quality or language.",
            }

        # 2) Analyze sentiment using audio + transcript
        result = analyze_sentiment_from_audio(audio_path, transcript)

        # 3) Classify scenario based on transcript
        scenario_info = classify_scenario_from_transcript(transcript)

        # 4) Normalise score (0–100)
        raw_score = result.get("score")
        score = normalize_score(raw_score)

        # 5) Save to Excel
        save_result_to_excel(result, audio_path, transcript, scenario_info)

        # 6) Return everything needed for UI
        return {
            "success": True,
            "audio_path": audio_path,
            "sentiment": result.get("sentiment"),
            "score": score,  # already 0–100 integer
            "tone": result.get("tone"),
            "explanation": result.get("explanation"),
            "scenario_id": scenario_info.get("scenario_id"),
            "scenario_title": scenario_info.get("scenario_title"),
            "transcript": transcript,
            "english": english,
        }

    except SystemExit:
        # Quota error or sys.exit from safe_generate_content
        return {
            "success": False,
            "error": "Quota exceeded or system exit triggered. Check scheduler_log.txt.",
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
        }
def analyze_single_text_for_ui(text: str, source_name: str = "document") -> dict:
    """
    UI helper for Flask (TEXT ONLY).

    It:
      - (optional) formalises text if ENABLE_FORMALISATION is True
      - runs sentiment using text-only mode
      - classifies scenario
      - saves to Excel
      - returns all important values in a dict for the web UI
    """
    try:
        transcript = (text or "").strip()
        if not transcript:
            return {"success": False, "error": "Document text is empty."}

        # Optional formalisation (reuse your existing cleaner)
        if ENABLE_FORMALISATION:
            try:
                transcript = formalise_transcript(transcript).strip()
            except Exception:
                pass  # if formalisation fails, keep original

        # 1) Sentiment from text (Gemini text-only)
        result = analyze_sentiment(transcript)

        # 2) Scenario classification from transcript
        scenario_info = classify_scenario_from_transcript(transcript)

        # 3) Normalise score
        raw_score = result.get("score")
        score = normalize_score(raw_score)

        # 4) Save to Excel (document results)
        # We reuse save_result_to_excel, but it expects an "audio_path".
        # We'll pass a fake filename so Excel "Audio File" column stores the document name.
        fake_audio_path = source_name  # e.g. "Client2.pdf"
        save_result_to_excel(result, fake_audio_path, transcript, scenario_info)

        return {
            "success": True,
            "audio_path": source_name,
            "sentiment": result.get("sentiment"),
            "score": score,
            "tone": result.get("tone"),
            "explanation": result.get("explanation"),
            "scenario_id": scenario_info.get("scenario_id"),
            "scenario_title": scenario_info.get("scenario_title"),
            "transcript": transcript,
            "english": "",
        }

    except Exception as e:
        return {"success": False, "error": str(e)}



# ========================
# CLI helpers
# ========================
def process_single_audio_file(audio_path: str):
    """Helper: transcribe 1 audio file and analyze sentiment using tone."""
    try:
        # 1) Transcribe audio and get ONE transcript (formal if available)
        trans_result = transcribe_audio_to_text(audio_path)
        transcript = (trans_result.get("transcript") or "").strip()
        english = (trans_result.get("english") or "").strip() if trans_result.get("english") else ""

        # Show the single (formal) transcript
        print("\n--- Transcript (Formal) ---")
        print(transcript if transcript else "[Empty transcript]")

        # Optional: show English translation only if you enabled it
        if ENABLE_TRANSLATION:
            print("\n--- English Translation ---")
            print(english if english else "[Empty translation]")

        if not transcript:
            print("[Warning] Transcript is empty, skipping sentiment analysis.")
            return

        # 2) Analyze sentiment using AUDIO (tone) + the (formal) transcript
        result = analyze_sentiment_from_audio(audio_path, transcript)

        # 3) Classify scenario based on the transcript
        scenario_info = classify_scenario_from_transcript(transcript)

        print("\n--- Result (Audio with Tone) ---")
        print(f"File      : {audio_path}")
        print(f"Sentiment : {result.get('sentiment')}")

        raw_score = result.get("score")
        score = normalize_score(raw_score)
        print(f"Score     : {score}%")

        print(f"Tone      : {result.get('tone')}")
        print(f"Reason    : {result.get('explanation')}")
        print(f"Scenario  : ID {scenario_info.get('scenario_id')} - {scenario_info.get('scenario_title')}")
        print("-" * 50)

        # Save result into Excel
        save_result_to_excel(result, audio_path, transcript, scenario_info)

    except SystemExit:
        # Quota error handled in safe_generate_content; just propagate
        raise
    except Exception as e:
        print(f"\n[Error while processing {audio_path}]")
        print(e)
        print("-" * 50)


def process_all_audio_in_folder(folder: str):
    """
    Process ALL .wav files inside a given folder.
    """
    if not os.path.isdir(folder):
        print(f"\n[Error] Folder not found: {folder}")
        return

    files = [f for f in os.listdir(folder) if f.lower().endswith(".wav")]

    if not files:
        print(f"\n[Info] No .wav files found in folder: {folder}")
        return

    print(f"\n[Info] Found {len(files)} audio file(s) in folder: {folder}")
    for name in files:
        print(" -", name)

    for name in files:
        full_path = os.path.join(folder, name)
        process_single_audio_file(full_path)


# ========================
# CLI entry point
# ========================
def main():
    """
    Two modes:
    - Normal (interactive): python main.py
    - Auto mode:           python main.py --auto <folder_path>
    """
    args = sys.argv[1:]

    # --- AUTO MODE ---
    if len(args) >= 1 and args[0] == "--auto":
        folder = "audio_inputs"
        if len(args) >= 2:
            folder = args[1]

        print(f"[Auto mode] Processing folder: {folder}")

        try:
            with open("scheduler_log.txt", "a", encoding="utf-8") as log:
                log.write(f"Auto mode started for folder: {folder}\n")
        except Exception:
            pass

        try:
            process_all_audio_in_folder(folder)
            with open("scheduler_log.txt", "a", encoding="utf-8") as log:
                log.write("Auto mode completed successfully.\n")
        except SystemExit:
            return
        except Exception as e:
            try:
                with open("scheduler_log.txt", "a", encoding="utf-8") as log:
                    log.write(f"Auto mode FAILED: {str(e)}\n")
            except Exception:
                pass
        return

    # --- NORMAL INTERACTIVE MODE ---
    print("=== Gemini Sentiment Analysis Client (Text + Audio) ===")
    print("Options:")
    print("  1 - Type text feedback")
    print("  2 - Analyze ONE audio file (.wav)")
    print("  3 - Analyze ALL .wav files in default folder (audio_inputs)")
    print("  4 - Analyze all .wav files by DATE folder (e.g. audio_inputs/2025-12-01)")
    print("  q - Quit")

    while True:
        choice = input("\nSelect option (1/2/3/4/q): ").strip().lower()

        if choice in {"q", "quit", "exit"}:
            print("\nGoodbye!")
            break

        elif choice == "1":
            user_input = input("\nType client feedback text: ").strip()
            if not user_input:
                print("Please type something.")
                continue

            try:
                result = analyze_sentiment(user_input)
                print("\n--- Result (Text) ---")
                print(f"Sentiment : {result.get('sentiment')}")
                raw_score = result.get("score")
                score = normalize_score(raw_score)
                print(f"Score     : {score}%")
                print(f"Tone      : {result.get('tone')}")
                print(f"Reason    : {result.get('explanation')}")
            except SystemExit:
                break
            except Exception as e:
                print("\n[Error while analyzing text]")
                print(e)

        elif choice == "2":
            print("\nMake sure your audio file is in .wav format.")
            audio_path = input("Enter path to .wav file (e.g. audio_inputs/client1.wav): ").strip()

            if not audio_path:
                print("Please enter a valid file path.")
                continue

            try:
                process_single_audio_file(audio_path)
            except SystemExit:
                break

        elif choice == "3":
            folder = "audio_inputs"
            try:
                process_all_audio_in_folder(folder)
            except SystemExit:
                break

        elif choice == "4":
            date_str = input("Enter date folder name (e.g. 2025-12-01): ").strip()
            if not date_str:
                print("Please enter a valid date folder name.")
                continue

            folder = os.path.join("audio_inputs", date_str)
            print(f"\n[Info] Processing folder: {folder}")
            try:
                process_all_audio_in_folder(folder)
            except SystemExit:
                break

        else:
            print("Invalid choice. Please select 1, 2, 3, 4 or q.")


if __name__ == "__main__":
    main()
