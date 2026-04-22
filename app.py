import sys, os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

import os
import re
import json
from datetime import datetime
from functools import wraps
from io import BytesIO
from docx import Document
from pypdf import PdfReader
from db import get_connection
from werkzeug.security import generate_password_hash, check_password_hash
from pywebpush import webpush, WebPushException
from services.dashboard_service import build_dashboard_data
from dotenv import load_dotenv

load_dotenv()

import logging
from logging.handlers import TimedRotatingFileHandler

from flask import (
    Flask,
    Response,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    session,
    send_file,
    abort,
    jsonify,
)

from openpyxl import load_workbook,Workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import uuid
from threading import Thread, Lock

print("RUNNING:", __file__)

JOBS = {}          # job_id -> {"status": "...", "message": "...", "username": "..."}
JOBS_LOCK = Lock() # avoid race issues when multiple jobs update JOBS
PUSH_SUBSCRIPTIONS = {}   # username -> push subscription JSON
PUSH_SUB_FILE = os.path.join(os.getcwd(), "push_subscriptions.json")
PUSH_LOCK = Lock()

# Import from main.py
from main import analyze_single_audio_for_ui, analyze_single_text_for_ui, EXCEL_PATH, TRANSCRIPT_DIR
PROJECT_ROOT = os.path.dirname(EXCEL_PATH)
NOTIF_SHEET_NAME = "Notifications"

# ========================
# Admin User Account helpers
# ========================

# -----------------------------------------------------------------------------
# FUNCTION: upsert_user_account
# PURPOSE: "Upsert" means "Update or Insert". This checks if a user exists in 
#          the SQL database. If they do, it updates their role. If they don't, 
#          it creates a brand new account for them with a secure hashed password.
# -----------------------------------------------------------------------------
def upsert_user_account(username: str, role: str, password_plain: str = ""):
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            IF EXISTS (SELECT 1 FROM dbo.user_account WHERE username = ?)
                UPDATE dbo.user_account
                SET role = ?
                WHERE username = ?
            ELSE
                INSERT INTO dbo.user_account (username, full_name, email, role, user_password)
                VALUES (?, ?, ?, ?, ?)
            """,
            (
                username,
                role,
                username,
                username,            # for INSERT username
                username,            # full_name default
                "",                  # email default
                role,
                generate_password_hash(password_plain) if password_plain else generate_password_hash("temp12345"),
            )
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()

# -----------------------------------------------------------------------------
# FUNCTION: create_user_account
# PURPOSE: Strictly used to create a new user in the SQL database. It takes a 
#          plain text password, hashes it for security, and saves the record.
# -----------------------------------------------------------------------------
def create_user_account(username: str, full_name: str, email: str, role: str, password_plain: str):
    conn = get_connection()
    cur = conn.cursor()
    try:
        password_hash = generate_password_hash(password_plain)

        cur.execute(
            """
            INSERT INTO dbo.user_account
              (username, full_name, email, role, user_password)
            VALUES (?, ?, ?, ?, ?)
            """,
            (username, full_name, email, role, password_hash)
        )

        conn.commit()
        return cur.lastrowid
    finally:
        cur.close()
        conn.close()

# -----------------------------------------------------------------------------
# FUNCTION: fetch_all_user_account
# PURPOSE: Grabs a list of all users from the SQL database. If a search query (q) 
#          is provided, it filters the results by username, name, or ID. Used 
#          for the Admin Dashboard to manage staff.
# -----------------------------------------------------------------------------
def fetch_all_user_account(q: str = ""):
    conn = get_connection()
    cur = conn.cursor()
    try:
        if q:
            like = f"%{q}%"
            cur.execute(
                """
                SELECT
                    userID AS id,
                    username,
                    full_name,
                    email,
                    role,
                    created_at
                FROM dbo.user_account
                WHERE username LIKE ?
                   OR full_name LIKE ?
                   OR CAST(userID AS VARCHAR(50)) LIKE ?
                ORDER BY userID DESC
                """,
                (like, like, like)
            )
        else:
            cur.execute(
                """
                SELECT
                    userID AS id,
                    username,
                    full_name,
                    email,
                    role,
                    created_at
                FROM dbo.user_account
                ORDER BY userID DESC
                """
            )
        rows = cur.fetchall()
        return rows_to_dict_list(cur, rows)
    finally:
        cur.close()
        conn.close()

# -----------------------------------------------------------------------------
# FUNCTION: fetch_user_by_username
# PURPOSE: Looks up a specific user in the database by their username. This is 
#          the core function used during the Login process to check passwords.
# -----------------------------------------------------------------------------
def fetch_user_by_username(username: str):
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            SELECT 
                userID AS id,
                username,
                full_name,
                email,
                role,
                user_password AS password_hash
            FROM dbo.user_account
            WHERE username=?
            """,
            (username,)
        )
        row = cur.fetchone()
        return row_to_dict(cur, row)
    finally:
        cur.close()
        conn.close()

# -----------------------------------------------------------------------------
# FUNCTION: username_or_email_exists
# PURPOSE: A validation check. Before creating a new account, it ensures the 
#          username or email isn't already taken by someone else in the DB.
# -----------------------------------------------------------------------------
def username_or_email_exists(username: str, email: str):
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            SELECT
              (SELECT COUNT(*) FROM dbo.user_account WHERE username=?) AS u_count,
              (SELECT COUNT(*) FROM dbo.user_account WHERE email=?) AS e_count
            """,
            (username, email)
        )
        row = row_to_dict(cur, cur.fetchone()) or {}
        return (row.get("u_count", 0) > 0), (row.get("e_count", 0) > 0)
    finally:
        cur.close()
        conn.close()

# -----------------------------------------------------------------------------
# FUNCTION: fetch_user_account_by_id
# PURPOSE: Finds a specific user using their internal Database ID instead of 
#          their username. Used when editing an account from the Admin panel.
# -----------------------------------------------------------------------------
def fetch_user_account_by_id(user_id: int):
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            SELECT
                userID AS id,
                username,
                full_name,
                email,
                role
            FROM dbo.user_account
            WHERE userID = ?
            """,
            (user_id,)
        )
        row = cur.fetchone()
        return row_to_dict(cur, row)
    finally:
        cur.close()
        conn.close()

# -----------------------------------------------------------------------------
# FUNCTION: update_user_account
# PURPOSE: Saves changes (like a new name, email, or role) made to an existing 
#          user account back into the SQL database.
# -----------------------------------------------------------------------------
def update_user_account(user_id: int, full_name: str, email: str, role: str):
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            UPDATE dbo.user_account
            SET full_name = ?, email = ?, role = ?
            WHERE userID = ?
            """,
            (full_name, email, role, user_id)
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()

# -----------------------------------------------------------------------------
# FUNCTION: delete_user_account
# PURPOSE: Completely removes a user account from the SQL database.
# -----------------------------------------------------------------------------
def delete_user_account(user_id: int):
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            "DELETE FROM dbo.user_account WHERE userID = ?",
            (user_id,)
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()

# =========================
# Scenario Master (Excel) - load ALL Scenario IDs
# =========================
SCENARIO_MASTER_FILE = os.path.join(os.getcwd(), "Scenario EIMY.xlsx")

# -----------------------------------------------------------------------------
# FUNCTION: load_all_scenario_ids_from_excel
# PURPOSE: Opens the master Eastspring Excel file and extracts every single 
#          possible Scenario ID (e.g., 'ME-01', 'ME-02'). This helps populate 
#          dropdowns and charts so they show all categories, even empty ones.
# -----------------------------------------------------------------------------
def load_all_scenario_ids_from_excel():
    """
    Reads ALL Scenario IDs from Eastspring_Scenarios.xlsx.
    Expected header contains 'Scenario ID' (case-insensitive).
    Returns list like [1,2,3,...] or ['ME-01','ME-02',...]
    """
    try:
        wb = load_workbook(SCENARIO_MASTER_FILE, data_only=True)
        ws = wb.active  

        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).strip().lower() if cell.value else "")

        #  column that contains "scenario id"
        if "scenario id" not in headers:
            print("⚠️ Scenario ID column not found in Eastspring_Scenarios.xlsx")
            return []

        col_idx = headers.index("scenario id") + 1

        ids = []
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is None:
                continue
            v = str(v).strip()
            if not v:
                continue
            ids.append(v)

        # remove duplicates but keep order
        seen = set()
        out = []
        for x in ids:
            if x not in seen:
                seen.add(x)
                out.append(x)

        return out
    except Exception as e:
        print("⚠️ Failed to load Scenario IDs:", e)
        return []

ALL_SCENARIO_IDS = load_all_scenario_ids_from_excel()

# ========================
# Helpers
# ======================== 

# -----------------------------------------------------------------------------
# FUNCTION: load_push_subs
# PURPOSE: Reads the local JSON file that stores user browser subscriptions. 
#          This is how the server knows which browser to send web push notifications to.
# -----------------------------------------------------------------------------
def load_push_subs():
    if not os.path.exists(PUSH_SUB_FILE):
        return {}
    try:
        with open(PUSH_SUB_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

# -----------------------------------------------------------------------------
# FUNCTION: save_push_subs
# PURPOSE: Writes new browser subscriptions to the JSON file so they are 
#          remembered even if the Flask server restarts.
# -----------------------------------------------------------------------------
def save_push_subs(data):
    with open(PUSH_SUB_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f)

PUSH_SUBSCRIPTIONS = load_push_subs()
 
# -----------------------------------------------------------------------------
# FUNCTION: parse_date
# PURPOSE: Converts a standard HTML string date (YYYY-MM-DD) from a webpage form 
#          into a true Python Date object so it can be used for filtering data.
# -----------------------------------------------------------------------------
def parse_date(d):
    # HTML <input type="date"> returns YYYY-MM-DD
    if not d:
        return None
    try:
        return datetime.strptime(d, "%Y-%m-%d").date()
    except ValueError:
        return None

# -----------------------------------------------------------------------------
# FUNCTION: detect_file_type
# PURPOSE: Looks at a filename extension (.wav, .pdf) and returns a clean string 
#          representing the type. Used to show the correct icon on the web dashboard.
# -----------------------------------------------------------------------------
def detect_file_type(filename: str) -> str:
    name = (filename or "").lower()
    if name.endswith(".wav"):
        return "wav"
    if name.endswith(".pdf"):
        return "pdf"
    if name.endswith(".docx"):
        return "docx"
    if name.endswith(".xlsx") or name.endswith(".xls"):
        return "excel"
    return "unknown"

# -----------------------------------------------------------------------------
# FUNCTION: detect_source_type
# PURPOSE: Groups file types into broader categories ("audio" vs "text"). 
#          Important for the AI to know if it's listening to a voice or reading a doc.
# -----------------------------------------------------------------------------
def detect_source_type(filename: str) -> str:
    ext = detect_file_type(filename)
    if ext == "wav":
        return "audio"
    if ext in ("pdf", "docx"):
        return "text"
    return "unknown"

# -----------------------------------------------------------------------------
# FUNCTION: format_dt
# PURPOSE: Takes a messy Python datetime object and turns it into a clean, 
#          human-readable string (e.g., "15 Mar 2026 \n 02:30 PM").
# -----------------------------------------------------------------------------
def format_dt(dt):
    if not dt:
        return "Not available"

    if not isinstance(dt, datetime):
        try:
            dt = datetime.fromisoformat(str(dt))
        except Exception:
            return str(dt)

    date_part = dt.strftime("%d %b %Y")
    time_part = dt.strftime("%I:%M %p")
    return f"{date_part}\n{time_part}"

# -----------------------------------------------------------------------------
# FUNCTION: allowed_file
# PURPOSE: Security check! Ensures users only upload .wav files when submitting audio.
# -----------------------------------------------------------------------------
def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in {"wav"}

# -----------------------------------------------------------------------------
# FUNCTION: allowed_doc_file
# PURPOSE: Security check! Ensures users only upload .pdf or .docx when submitting text.
# -----------------------------------------------------------------------------
def allowed_doc_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in {"pdf", "docx"}

# -----------------------------------------------------------------------------
# FUNCTION: format_dt_parts
# PURPOSE: Similar to format_dt, but splits the date and time into a tuple 
#          (two separate strings) so the UI can style them differently.
# -----------------------------------------------------------------------------
def format_dt_parts(dt):
    """Return date and time as two strings (date on one line, time on next)."""
    if not dt:
        return ("Not available", "")

    if not isinstance(dt, datetime):
        try:
            dt = datetime.fromisoformat(str(dt))
        except Exception:
            return (str(dt), "")

    date_part = dt.strftime("%d %b %Y")
    time_part = dt.strftime("%I:%M %p")
    return (date_part, time_part)

# -----------------------------------------------------------------------------
# FUNCTION: run_upload_analysis_job
# PURPOSE: The Background Worker Engine! This massive function runs silently in 
#          the background so the user's web page doesn't freeze during heavy uploads.
#          It processes PDFs and Audio, talks to the AI, saves to the DB, and sends
#          a push notification when the job is completely done.
# -----------------------------------------------------------------------------
def run_upload_analysis_job(job_id: str, username: str, doc_files_meta: list, audio_paths: list):
    app.logger.info(f"[JOB {job_id}] START user={username} docs={len(doc_files_meta)} audios={len(audio_paths)}")

    success_count = 0
    failure_count = 0

    try:
        with JOBS_LOCK:
            JOBS[job_id] = {"status": "running", "message": "Processing...", "username": username}

        # --- documents ---
        for idx, item in enumerate(doc_files_meta, start=1):
            filename = item["filename"]
            ext = item["ext"]
            data_bytes = item["bytes"]

            app.logger.info(f"[JOB {job_id}] DOC {idx}/{len(doc_files_meta)} START {filename}")

            try:
                from io import BytesIO
                fs = BytesIO(data_bytes)

                class _Tmp:
                    def read(self):
                        return fs.getvalue()

                text = extract_text_from_pdf(_Tmp()) if ext == "pdf" else extract_text_from_docx(_Tmp())
                if not (text or "").strip():
                    app.logger.warning(f"[JOB {job_id}] DOC SKIP (no text) {filename}")
                    failure_count += 1
                    continue

                data = analyze_single_text_for_ui(text, os.path.basename(filename))
                app.logger.info(f"[JOB {job_id}] DOC DONE {filename} success={data.get('success')} sentiment={data.get('sentiment')}")

                if not data.get("success"):
                    failure_count += 1
                    continue

                saved_ok = save_sentiment_to_db(
                    username=username,
                    source_type="text",
                    filename=os.path.basename(filename),
                    text_input=text,
                    sentiment=data.get("sentiment"),
                    confidence=data.get("score"),
                    tone=data.get("tone"),
                    scenario_id=data.get("scenario_id"),
                    scenario_title=data.get("scenario_title"),
                    explanation=data.get("explanation"),
                    transcript=data.get("transcript"),
                )

                if not saved_ok:
                    app.logger.error(f"[JOB {job_id}] DOC DB SAVE FAIL {filename}")
                    failure_count += 1
                    continue

                create_notification(
                    username=username,
                    message=f"Document analyzed: {os.path.basename(filename)} ({data.get('sentiment')})",
                    url="/sentiment_result"
                )

                success_count += 1

            except Exception as e:
                app.logger.error(f"[JOB {job_id}] DOC ERROR {filename}: {e}")
                failure_count += 1
                continue

        # --- audios ---
        for idx, file_path in enumerate(audio_paths, start=1):
            filename = os.path.basename(file_path)
            app.logger.info(f"[JOB {job_id}] AUDIO {idx}/{len(audio_paths)} START {filename}")

            try:
                label, score, details, raw_data = analyze_sentiment_from_wav(file_path)
                app.logger.info(f"[JOB {job_id}] AUDIO DONE {filename} label={label} score={score}")

                scenario_id = raw_data.get("scenario_id")
                scenario_title = raw_data.get("scenario_title")
                transcript = raw_data.get("transcript") or ""

                text_input_value = f"Scenario ID {scenario_id} - {scenario_title}\n\n{transcript}".strip()

                saved_ok = save_sentiment_to_db(
                    username=username,
                    source_type="audio",
                    filename=os.path.basename(filename),
                    text_input=text_input_value,
                    sentiment=label,
                    confidence=score,
                    tone=raw_data.get("tone"),
                    scenario_id=scenario_id,
                    scenario_title=scenario_title,
                    explanation=raw_data.get("explanation"),
                    transcript=transcript,
                )

                if not saved_ok:
                    app.logger.error(f"[JOB {job_id}] AUDIO DB SAVE FAIL {filename}")
                    failure_count += 1
                    continue

                create_notification(
                    username=username,
                    message=f"Audio analyzed: {filename} ({label})",
                    url="/sentiment_result"
                )

                transcript_text = (raw_data.get("transcript") or "").strip()
                if transcript_text:
                    safe_name = re.sub(r'[<>:"/\\|?*]', "_", filename)
                    safe_base = os.path.splitext(safe_name)[0]
                    txt_path = os.path.join(TRANSCRIPT_FOLDER, f"{safe_base}.txt")
                    with open(txt_path, "w", encoding="utf-8") as f:
                        f.write(transcript_text)

                success_count += 1

            except Exception as e:
                app.logger.error(f"[JOB {job_id}] AUDIO ERROR {filename}: {e}")
                failure_count += 1
                continue

        # --- final job status ---
        if success_count == 0:
            with JOBS_LOCK:
                JOBS[job_id] = {
                    "status": "error",
                    "message": f"No files processed successfully. Failed: {failure_count}",
                    "username": username
                }
            app.logger.error(f"[JOB {job_id}] MARKED ERROR ❌ success={success_count} failure={failure_count}")
            return

        with JOBS_LOCK:
            JOBS[job_id] = {
                "status": "done",
                "message": f"Sentiment result is ready ✅ ({success_count} success, {failure_count} failed)",
                "username": username
            }

        app.logger.info(f"[JOB {job_id}] MARKED DONE ✅ success={success_count} failure={failure_count}")

        ok, msg = send_push_to_user(
            username,
            "Sentiment ready ✅",
            f"Your sentiment result is ready. ({success_count} success, {failure_count} failed)"
        )
        app.logger.info(f"[JOB {job_id}] PUSH: ok={ok} msg={msg}")

    except Exception as e:
        app.logger.error(f"[JOB {job_id}] FATAL ERROR: {e}")
        with JOBS_LOCK:
            JOBS[job_id] = {"status": "error", "message": str(e), "username": username}

        try:
           send_push_to_user(username, "Analysis failed ❌", str(e))
        except Exception as e2:
           app.logger.error(f"[JOB {job_id}] PUSH FAIL NOTICE ERROR (ignored): {e2}")

# -----------------------------------------------------------------------------
# FUNCTION: save_sentiment_to_db
# PURPOSE: Takes all the answers the AI just figured out (sentiment, tone, score) 
#          and writes them permanently into the `dbo.sentiments` SQL Server table.
# -----------------------------------------------------------------------------
def save_sentiment_to_db(
    username,
    source_type,
    filename,
    text_input,
    sentiment,
    confidence,
    tone=None,
    scenario_id=None,
    scenario_title=None,
    explanation=None,
    transcript=None,
):
    conn = None
    cur = None
    try:
        conn = get_connection()
        cur = conn.cursor()
        
        # We grab the exact local time right now
        local_time_now = datetime.now() 
        
        cur.execute(
            """
            INSERT INTO sentiments
            (username, source_type, filename, text_input, sentiment, confidence,
             tone, scenario_id, scenario_title, explanation, transcript, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                username, source_type, filename, text_input, sentiment, confidence,
                tone, scenario_id, scenario_title, explanation, transcript, local_time_now
            )
        )
        conn.commit()
        app.logger.info(f"[DB OK] Inserted into dbo.sentiments: {filename} {sentiment} {source_type}")
        return True
    
    except Exception as e: 
        app.logger.error(f"[DB FAIL] Insert into dbo.sentiments failed: {repr(e)}")
        return False
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()

# -----------------------------------------------------------------------------
# FUNCTION: extract_id_only
# PURPOSE: Cleans up messy text from the AI. If the AI outputs "ID ME-07 - App Issue", 
#          this function strips it down to purely "ME-07" for the database.
# -----------------------------------------------------------------------------
def extract_id_only(value):
    """
    Converts:
    'ID 8 - Feedback / Complaint about Service' -> '8'
    'ID ME-07 - App Performance Issue' -> 'ME-07'
    'ME-08' -> 'ME-08'
    """
    v = str(value).strip()

    # remove leading "ID "
    if v.upper().startswith("ID "):
        v = v[3:].strip()

    # cut off anything after " - "
    if " - " in v:
        v = v.split(" - ", 1)[0].strip()

    return v

# -----------------------------------------------------------------------------
# FUNCTION: scenario_sort_key
# PURPOSE: A specialized sorting rule. It ensures that when charts display 
#          Scenario IDs, they are ordered logically (ME-01, ME-02, 1, 2) instead 
#          of random alphabetical order.
# -----------------------------------------------------------------------------
def scenario_sort_key(sid: str):
    """
    Order rule:
    1) ME-xx first (numeric order)
    2) Pure numbers next (1,2,3...)
    3) Other prefixes (INV, etc.)
    """
    s = str(sid).strip().upper()

    # ME-xx
    m = re.match(r"^ME[-_]?(\d+)$", s)
    if m:
        return (0, int(m.group(1)))

    # numeric only
    if s.isdigit():
        return (1, int(s))

    # other prefix-number (INV-011, etc.)
    m = re.match(r"^([A-Z]+)[-_]?(\d+)$", s)
    if m:
        return (2, m.group(1), int(m.group(2)))

    # fallback
    return (3, s)

# -----------------------------------------------------------------------------
# FUNCTION: get_yearly_sentiment_overview
# PURPOSE: Queries the SQL database to count how many Complaints vs Non-Complaints 
#          happened in each month of the selected year. Powers the Line Chart.
# -----------------------------------------------------------------------------
def get_yearly_sentiment_overview(username, selected_year, source_type=""):
    """
    Returns:
      line_complaint[12], line_non[12]
    using SQL Server sentiments table filtered by username/year/(optional source_type).
    """
    conn = get_connection()
    cur = conn.cursor()

    sql = """
        SELECT MONTH(created_at) AS m, sentiment, COUNT(*) AS total
        FROM sentiments
        WHERE username = ?
          AND YEAR(created_at) = ?
    """
    params = [username, selected_year]

    if source_type in ("audio", "text"):
        sql += " AND source_type = ?"
        params.append(source_type)

    sql += " GROUP BY MONTH(created_at), sentiment"

    cur.execute(sql, params)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    complaint = [0] * 12
    non = [0] * 12

    for m, sentiment, total in rows:
        idx = int(m) - 1  # 0..11
        s = normalize_sentiment(sentiment)
        if s == "complaint":
            complaint[idx] += int(total)
        elif s == "non":
            non[idx] += int(total)

    return complaint, non

# -----------------------------------------------------------------------------
# FUNCTION: get_scenario_overview
# PURPOSE: Queries the SQL database to count how many issues hit specific Scenario 
#          IDs (like ME-07) in a given month. Powers the Scenario Bar Chart.
# -----------------------------------------------------------------------------
def get_scenario_overview(username, selected_year, selected_month, source_type=""):
    """
    Returns scenario_labels (ID only), scenario_complaint, scenario_non
    - Includes ALL Scenario IDs from master Excel (even zero)
    - Counts from SQL Server sentiments table
    - Sorted: numeric IDs first, then ME-xx (and other prefix-xx) in correct order
    """
    conn = get_connection()
    cur = conn.cursor()

    sql = """
        SELECT scenario_id, sentiment, COUNT(*) as total
        FROM sentiments
        WHERE username = ?
          AND YEAR(created_at) = ?
          AND MONTH(created_at) = ?
    """
    params = [username, selected_year, selected_month]

    if source_type in ("audio", "text"):
        sql += " AND source_type = ?"
        params.append(source_type)

    sql += " GROUP BY scenario_id, sentiment"

    cur.execute(sql, params)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    # 1) Start with ALL scenario IDs from Excel as 0
    master_ids = [extract_id_only(x) for x in (ALL_SCENARIO_IDS or []) if str(x).strip()]
    scenario_map = {sid: {"complaint": 0, "non": 0} for sid in master_ids}

    # 2) Fill real DB counts
    for scenario_id, sentiment, total in rows:
        sid = extract_id_only(scenario_id) if scenario_id is not None else "UNKNOWN"

        if sid not in scenario_map:
            scenario_map[sid] = {"complaint": 0, "non": 0}

        s = normalize_sentiment(sentiment)
        if s == "complaint":
            scenario_map[sid]["complaint"] += int(total)
        elif s == "non":
            scenario_map[sid]["non"] += int(total)

    # 3) Sort IDs (numeric first, then ME-xx correctly, etc.)
    ordered_ids = sorted(scenario_map.keys(), key=scenario_sort_key)

    # Labels: ID only (no title)
    scenario_labels = ordered_ids
    scenario_c = [scenario_map[sid]["complaint"] for sid in ordered_ids]
    scenario_n = [scenario_map[sid]["non"] for sid in ordered_ids]

    return scenario_labels, scenario_c, scenario_n

# ========================
# Push Notification Helper 
# ========================

# -----------------------------------------------------------------------------
# FUNCTION: send_push_to_user
# PURPOSE: Uses web push technology to send a pop-up alert to the user's browser 
#          (e.g., "Analysis Complete!") even if they are on a different tab.
# -----------------------------------------------------------------------------
def send_push_to_user(username: str, title: str, body: str):
    if not (VAPID_PUBLIC_KEY and VAPID_PRIVATE_KEY_PATH):
        return False, "VAPID keys not configured"

    with PUSH_LOCK:
        sub = PUSH_SUBSCRIPTIONS.get(username)

    if not sub:
        return False, "No push subscription saved for this user"

    payload = json.dumps({
    "title": title,
    "body": body,
    "url": "/sentiment_result"
    })

    try:
        webpush(
            subscription_info=sub,
            data=payload,
            vapid_private_key=VAPID_PRIVATE_KEY_PATH,
            vapid_claims=VAPID_CLAIMS,
        )
        return True, "Sent"
    except WebPushException as e:
        return False, f"Push failed: {repr(e)}"
    except Exception as e:
        return False, f"Push failed: {e}"

# ========================
# Document text extraction
# ========================

# -----------------------------------------------------------------------------
# FUNCTION: extract_text_from_docx
# PURPOSE: Opens a Microsoft Word document (.docx) uploaded by the user and 
#          extracts all the raw paragraph text out of it so the AI can read it.
# -----------------------------------------------------------------------------
def extract_text_from_docx(file_storage) -> str:
    data = file_storage.read()
    doc = Document(BytesIO(data))
    text = "\n".join(p.text for p in doc.paragraphs if p.text and p.text.strip())
    return text.strip()

# -----------------------------------------------------------------------------
# FUNCTION: extract_text_from_pdf
# PURPOSE: Opens a PDF document uploaded by the user and attempts to scrape 
#          all readable text off the pages for the AI. (Fails on scanned images).
# -----------------------------------------------------------------------------
def extract_text_from_pdf(file_storage) -> str:
    data = file_storage.read()
    reader = PdfReader(BytesIO(data))
    parts = []
    for page in reader.pages:
        t = page.extract_text() or ""
        if t.strip():
            parts.append(t)
    return "\n".join(parts).strip()

# -----------------------------------------------------------------------------
# FUNCTION: analyze_sentiment_from_wav
# PURPOSE: A bridge function. It receives a .wav file path, hands it over to 
#          `main.py` (which talks to Google Gemini AI), and formats the AI's 
#          complex JSON response into a simple label and score for the website.
# -----------------------------------------------------------------------------
def analyze_sentiment_from_wav(wav_path: str):
    """
    Wrapper for the web UI that calls main.py and returns
    (label, score, details, raw_data).
    """
    data = analyze_single_audio_for_ui(wav_path)

    if data is None:
        raise RuntimeError("analyze_single_audio_for_ui returned None")

    if not data.get("success"):
        raise RuntimeError(data.get("error", "Unknown error from analysis"))

    label = data.get("sentiment")
    score = data.get("score")
    tone = data.get("tone")
    scenario_id = data.get("scenario_id")
    scenario_title = data.get("scenario_title")
    explanation = data.get("explanation")

    details = (
        f"Tone: {tone} | "
        f"Scenario: ID {scenario_id} - {scenario_title}\n"
        f"Reason: {explanation}"
    )

    return label, score, details, data


# ========================
# Flask app config
# ========================
UPLOAD_FOLDER = os.path.join(os.getcwd(), "uploaded_audio")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
TRANSCRIPT_FOLDER = TRANSCRIPT_DIR
os.makedirs(TRANSCRIPT_FOLDER, exist_ok=True)

# -----------------------------------------------------------------------------
# FUNCTION: row_to_dict
# PURPOSE: Converts a single row returned by a raw SQL query into a Python 
#          Dictionary, mapping column names to values so the UI can easily read it.
# -----------------------------------------------------------------------------
def row_to_dict(cursor, row):
    if row is None:
        return None
    columns = [col[0] for col in cursor.description]
    return dict(zip(columns, row))

# -----------------------------------------------------------------------------
# FUNCTION: rows_to_dict_list
# PURPOSE: Does the exact same thing as `row_to_dict`, but loops through 
#          multiple rows (like when fetching a whole history table).
# -----------------------------------------------------------------------------
def rows_to_dict_list(cursor, rows):
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in rows]

app = Flask(__name__)
app.secret_key = "some-secret-key"  # change for production
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# --- LOGGING CONFIGURATION ---
LOG_FOLDER = os.path.join(os.getcwd(), "logs")
os.makedirs(LOG_FOLDER, exist_ok=True)

log_file_path = os.path.join(LOG_FOLDER, "system.log")

# Rotates the file every single day at exact midnight. 
# Keeps 30 days of history and automatically deletes older ones to save space!
file_handler = TimedRotatingFileHandler(
    log_file_path, 
    when="midnight", 
    interval=1, 
    backupCount=30,
    encoding="utf-8"
)

# Tells Python exactly how to format the date in the file name
file_handler.suffix = "%Y-%m-%d.log"

formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
file_handler.setLevel(logging.INFO)

app.logger.addHandler(file_handler)
app.logger.setLevel(logging.INFO)

werkzeug_logger = logging.getLogger('werkzeug')
werkzeug_logger.addHandler(file_handler)
werkzeug_logger.setLevel(logging.INFO)


# ========================
# Push Notification Config
# ========================

VAPID_PUBLIC_KEY = os.getenv("VAPID_PUBLIC_KEY", "")

# define variable first so it always exists
VAPID_PRIVATE_KEY_PATH = ""

pem_path = os.getenv("VAPID_PRIVATE_KEY_PEM_PATH", "").strip()
if pem_path:
    abs_pem_path = os.path.join(os.path.dirname(__file__), pem_path)
    if os.path.exists(abs_pem_path):
        VAPID_PRIVATE_KEY_PATH = abs_pem_path
        print("VAPID private key loaded:", VAPID_PRIVATE_KEY_PATH)
    else:
        print("VAPID private key file NOT FOUND:", abs_pem_path)
else:
    print("VAPID_PRIVATE_KEY_PEM_PATH not set in .env")

VAPID_SUBJECT = os.getenv("VAPID_SUBJECT", "mailto:admin@company.com")
VAPID_CLAIMS = {"sub": VAPID_SUBJECT}

# Simple hardcoded login (change as needed)
USERS = {
    "admin": {"password": "password123", "role": "ADMIN"},
    "user1": {"password": "user123", "role": "USER"},
    "user2": {"password": "user456", "role": "USER"},
}


USER_PROFILES = {
    "admin": {
        "name": "Admin User",
        "position": "Customer Service Supervisor",
        "status": "Active",
        "department": "Helpdesk",
        "email": "admin@company.com",
    }
}

# -----------------------------------------------------------------------------
# FUNCTION: sw
# PURPOSE: A web route that serves the Service Worker script (`sw.js`). This is 
#          the background script the browser uses to receive Push Notifications.
# -----------------------------------------------------------------------------
@app.get("/sw.js")
def sw():
    return app.send_static_file("sw.js")


# ========================
# Auth helpers
# ========================

# -----------------------------------------------------------------------------
# FUNCTION: login_required
# PURPOSE: A Python decorator. You put `@login_required` above any web route 
#          that should be locked down. If a user isn't logged in, it kicks them 
#          out and redirects them to the login screen.
# -----------------------------------------------------------------------------
def login_required(f):
    """Decorator to protect routes so only logged-in users can access."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

# -----------------------------------------------------------------------------
# FUNCTION: admin_required
# PURPOSE: A stricter Python decorator. It checks if the user is logged in AND 
#          if their specific role is "ADMIN". Used to lock down management pages.
# -----------------------------------------------------------------------------
def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))

        if session.get("role") != "ADMIN":
            flash("You are not allowed to access admin pages.")
            return redirect(url_for("home"))

        return f(*args, **kwargs)
    return decorated


# ========================
# History helpers
# ========================

# -----------------------------------------------------------------------------
# FUNCTION: ensure_notifications_sheet
# PURPOSE: Safely checks if the "Notifications" sheet exists in the Excel DB. 
#          If it doesn't, it creates it and adds the header columns automatically.
# -----------------------------------------------------------------------------
def ensure_notifications_sheet(wb):
    """Create Notifications sheet + header if not exists."""
    if NOTIF_SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(NOTIF_SHEET_NAME)
        ws.append(["id", "datetime", "username", "message", "url", "is_read"])
        return ws

    ws = wb[NOTIF_SHEET_NAME]

    # ensure header
    if ws.max_row == 1 and ws["A1"].value is None:
        ws.append(["id", "datetime", "username", "message", "url", "is_read"])

    return ws

# -----------------------------------------------------------------------------
# FUNCTION: create_notification
# PURPOSE: Appends a new line into the Notifications Excel sheet when the AI 
#          finishes analyzing a file, marking it as unread.
# -----------------------------------------------------------------------------
def create_notification(username: str, message: str, url: str = ""):
    """Write new notification row into Excel."""
    wb = load_workbook(EXCEL_PATH)
    ws = ensure_notifications_sheet(wb)

    # generate ID (simple incremental)
    last_id = 0
    if ws.max_row >= 2:
        try:
            last_id = int(ws.cell(row=ws.max_row, column=1).value or 0)
        except Exception:
            last_id = 0

    new_id = last_id + 1
    now = datetime.now()

    ws.append([new_id, now, username, message, url, False])
    wb.save(EXCEL_PATH)

# -----------------------------------------------------------------------------
# FUNCTION: load_notifications
# PURPOSE: Reads the notification history for a specific user from Excel so 
#          the little bell icon on the website can show their latest alerts.
# -----------------------------------------------------------------------------
def load_notifications(username: str, limit: int = 10):
    """Read latest notifications (newest first)."""
    wb = load_workbook(EXCEL_PATH, data_only=True)
    if NOTIF_SHEET_NAME not in wb.sheetnames:
        return []

    ws = wb[NOTIF_SHEET_NAME]
    items = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue

        nid, dt_val, user, msg, url, is_read = (list(row) + [None]*6)[:6]
        if str(user) != str(username):
            continue

        # format date
        if isinstance(dt_val, datetime):
            dt_str = dt_val.strftime("%d/%m/%Y %I:%M %p")
        else:
            dt_str = str(dt_val or "")

        items.append({
            "id": nid,
            "created_at": dt_str,
            "message": msg or "",
            "url": url or "",
            "is_read": bool(is_read),
        })

    # newest first
    items.reverse()
    return items[:limit]

# -----------------------------------------------------------------------------
# FUNCTION: unread_count
# PURPOSE: Calculates the little red number on the notification bell by counting 
#          how many rows in Excel for this user have `is_read` set to False.
# -----------------------------------------------------------------------------
def unread_count(username: str) -> int:
    wb = load_workbook(EXCEL_PATH, data_only=True)
    if NOTIF_SHEET_NAME not in wb.sheetnames:
        return 0

    ws = wb[NOTIF_SHEET_NAME]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        nid, dt_val, user, msg, url, is_read = (list(row) + [None]*6)[:6]
        if str(user) == str(username) and not bool(is_read):
            count += 1
    return count

# -----------------------------------------------------------------------------
# FUNCTION: mark_notification_read
# PURPOSE: When a user clicks a specific notification in the UI, this updates 
#          that row in Excel to mark it as read.
# -----------------------------------------------------------------------------
def mark_notification_read(username: str, notif_id: int):
    wb = load_workbook(EXCEL_PATH)
    ws = ensure_notifications_sheet(wb)

    for r in range(2, ws.max_row + 1):
        nid = ws.cell(row=r, column=1).value
        user = ws.cell(row=r, column=3).value
        if str(user) == str(username) and int(nid) == int(notif_id):
            ws.cell(row=r, column=6).value = True
            break

    wb.save(EXCEL_PATH)

# -----------------------------------------------------------------------------
# FUNCTION: mark_all_notifications_read
# PURPOSE: Updates every single notification row belonging to a user in Excel 
#          to be True (read). Used for a "Clear All" button.
# -----------------------------------------------------------------------------
def mark_all_notifications_read(username: str):
    wb = load_workbook(EXCEL_PATH)
    ws = ensure_notifications_sheet(wb)

    for r in range(2, ws.max_row + 1):
        user = ws.cell(row=r, column=3).value
        if str(user) == str(username):
            ws.cell(row=r, column=6).value = True

    wb.save(EXCEL_PATH)

# -----------------------------------------------------------------------------
# FUNCTION: load_history_rows
# PURPOSE: Reads the SQL Server Database, pulling ONLY the rows belonging to 
#          the person currently logged in.
# -----------------------------------------------------------------------------
def load_history_rows():
    current_user = session.get("username")
    rows = []
    conn = get_connection()
    cur = conn.cursor()
    
    try:
        # Pull ONLY records for the logged-in user!
        # (We added the 'comment' column back into this list!)
        cur.execute(
            """
            SELECT 
                id,
                created_at AS datetime,
                filename AS audio_file,
                sentiment,
                confidence AS score,
                tone,
                explanation,
                scenario_id,
                scenario_title,
                transcript,
                comment,
                source_type
            FROM dbo.sentiments
            WHERE username = ?
            ORDER BY created_at ASC
            """,
            (current_user,)
        )
        
        db_rows = rows_to_dict_list(cur, cur.fetchall())
        
        for r in db_rows:
            score = r.get("score")
            score_display = ""
            try:
                s = float(score) if score is not None else 0
                score_display = f"{s*100:.0f}%" if s <= 1 else f"{s:.0f}%"
            except Exception:
                score_display = str(score) if score is not None else ""

            explanation = r.get("explanation") or ""
            short_summary = re.sub(r"\s+", " ", str(explanation)).strip()
            if len(short_summary) > 90:
                short_summary = short_summary[:90] + "..."

            audio_file = r.get("audio_file") or ""
            file_type = detect_file_type(audio_file)

            rows.append({
                "id": r.get("id"),  # The real database ID!
                "datetime": r.get("datetime"),
                "audio_file": audio_file,
                "file_type": file_type,
                "summary": short_summary,
                "sentiment": r.get("sentiment") or "",
                "score": score,
                "score_display": score_display,
                "tone": r.get("tone") or "",
                "explanation": explanation,
                "scenario_id": r.get("scenario_id"),
                "scenario_title": r.get("scenario_title") or "",
                "transcript": r.get("transcript") or "",
                "comment": r.get("comment") or "", # NO LONGER HARDCODED!
                "source_type": r.get("source_type"),
            })
            
        return rows
        
    finally:
        if cur: cur.close()
        if conn: conn.close()

# -----------------------------------------------------------------------------
# FUNCTION: get_history_entry
# PURPOSE: Uses `load_history_rows` to pull all rows, but then isolates and 
#          returns just one specific row based on its index number.
# -----------------------------------------------------------------------------
def get_history_entry(row_id: int):
    """Get one row from Excel by row index (same as UI index)."""
    rows = load_history_rows()
    rows.reverse()  # match UI order (newest first)

    try:
        return rows[int(row_id)]
    except (IndexError, ValueError):
        return None

# -----------------------------------------------------------------------------
# FUNCTION: delete_history_entry
# PURPOSE: Uses the SQL 'id' to permanently delete a row from the database.
# -----------------------------------------------------------------------------
def delete_history_entry(row_id: int):
    # 1) Get the specific row based on what the user clicked
    row = get_history_entry(row_id)
    if not row:
        return
    
    # 2) Get the actual SQL Database ID (e.g., 1009, 1010)
    db_id = row.get("id")
    
    # 3) Delete it from SQL Server
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM dbo.sentiments WHERE id = ?", (db_id,))
        conn.commit()
    except Exception as e:
        print("[DB ERROR] Delete failed:", e)
    finally:
        cur.close()
        conn.close()

# -----------------------------------------------------------------------------
# FUNCTION: update_history_entry
# PURPOSE: Uses the SQL 'id' to save user comment edits directly to the database.
# -----------------------------------------------------------------------------
def update_history_entry(row_id: int, updates: dict):
    # 1) Get the specific row
    row = get_history_entry(row_id)
    if not row:
        return
        
    # 2) Get the actual SQL Database ID
    db_id = row.get("id")
    
    # 3) Update it in SQL Server
    conn = get_connection()
    cur = conn.cursor()
    try:
        # Update the summary (explanation column)
        if "summary" in updates:
            cur.execute("UPDATE dbo.sentiments SET explanation = ? WHERE id = ?", (updates["summary"], db_id))
            
        # Update the comment (comment column)
        if "comment" in updates:
            cur.execute("UPDATE dbo.sentiments SET comment = ? WHERE id = ?", (updates["comment"], db_id))
            
        conn.commit()
    except Exception as e:
        print("[DB ERROR] Update failed:", e)
    finally:
        cur.close()
        conn.close()

# -----------------------------------------------------------------------------
# FUNCTION: normalize_sentiment
# PURPOSE: Standardizes messy text. It converts strings like "Non-Complaint" or 
#          "non complaint" strictly into just "non" or "complaint" so math logic works.
# -----------------------------------------------------------------------------
def normalize_sentiment(s: str) -> str:
    s = (s or "").strip().lower()
    if "non" in s:
        return "non"
    if "complaint" in s:
        return "complaint"
    return ""

# -----------------------------------------------------------------------------
# FUNCTION: month_add
# PURPOSE: A math helper for dates. If you add 3 months to November (11), this 
#          safely loops it over to February (2) of the next year.
# -----------------------------------------------------------------------------
def month_add(year: int, month: int, delta: int):
    # month is 1..12
    m = month + delta
    y = year
    while m > 12:
        y += 1
        m -= 12
    while m < 1:
        y -= 1
        m += 12
    return y, m

# -----------------------------------------------------------------------------
# FUNCTION: month_key
# PURPOSE: Creates a simple tuple like (2026, 3) from a date. Used as an index key.
# -----------------------------------------------------------------------------
def month_key(dt: datetime):
    return (dt.year, dt.month)

# -----------------------------------------------------------------------------
# FUNCTION: build_month_series
# PURPOSE: Generates a list of all sequential months between two dates. Used for charts.
# -----------------------------------------------------------------------------
def build_month_series(rows, start_y, start_m, end_y, end_m):
    # Create all months between start and end
    months = []
    y, m = start_y, start_m
    while (y < end_y) or (y == end_y and m <= end_m):
        months.append((y, m))
        y, m = month_add(y, m, 1)
    return months

# -----------------------------------------------------------------------------
# FUNCTION: filter_rows_by_range
# PURPOSE: Takes massive history data and trims it down to a specific time window 
#          (e.g., "only show me the last 6 months") for accurate reporting.
# -----------------------------------------------------------------------------
def filter_rows_by_range(rows, months_back: int, year: int):
    """
    Filter rows into a rolling window ending at:
    - if year == current year: end = current month
    - else: end = Dec of that year
    """
    now = datetime.now()
    end_y = year
    end_m = now.month if year == now.year else 12

    # start month = end month - (months_back-1)
    start_y, start_m = month_add(end_y, end_m, -(months_back - 1))

    filtered = []
    for r in rows:
        dt = r.get("datetime")
        if not isinstance(dt, datetime):
            continue
        y, m = dt.year, dt.month
        # between start and end (month-based)
        if (y, m) < (start_y, start_m) or (y, m) > (end_y, end_m):
            continue
        filtered.append(r)

    return filtered, (start_y, start_m, end_y, end_m)

# -----------------------------------------------------------------------------
# FUNCTION: build_history_pdf
# PURPOSE: Dynamically generates a PDF document from scratch using a specific 
#          row of data. This powers the "Download PDF" buttons on the website.
# -----------------------------------------------------------------------------
def build_history_pdf(row: dict):
    """
    Build a PDF for a single history row (one client/result).
    Saves it to disk and returns the path.
    """
    # ----- SAFE FILENAME -----
    base_name = row.get("audio_file") or "history_item"
    base_name = os.path.splitext(os.path.basename(str(base_name)))[0]
    base_name = re.sub(r'[<>:"/\\|?*]', "_", base_name)
    pdf_filename = f"history_{base_name}.pdf"
    pdf_path = os.path.join(r"C:\Users\User\Downloads", pdf_filename)
    # --------------------------

    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    elements = []
    styles = getSampleStyleSheet()
    body = styles["BodyText"]
    body.wordWrap = "CJK"   # make long text wrap nicely

    # Title
    title = Paragraph("CLIENT RESULT", styles["Title"])
    elements.append(title)
    elements.append(Spacer(1, 12))

    def P(text):
        """Paragraph helper with basic newline handling."""
        return Paragraph(str(text).replace("\n", "<br/>"), body)

    data = [
        ["Field",          "Value"],
        ["Date / Time",    P(format_dt(row.get("datetime")))],
        ["File Name",      P(row.get("audio_file") or "")],
        ["Sentiment",      P(row.get("sentiment") or "")],
        ["Score (%)",      P(row.get("score") or "")],
        ["Tone",           P(row.get("tone") or "")],
        ["Scenario ID",    P(row.get("scenario_id") or "")],
        ["Scenario Title", P(row.get("scenario_title") or "")],
        ["Explanation",    P(row.get("explanation") or "")],
        ["Transcript",     P(row.get("transcript") or "")],
        ["Comment / Feedback", P(row.get("comment") or "")],

    ]

    table = Table(
        data,
        colWidths=[120, 600],
        repeatRows=1,
    )

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
    ]))

    elements.append(table)
    doc.build(elements)
    return pdf_path

# -----------------------------------------------------------------------------
# FUNCTION: run_analysis_job
# PURPOSE: Appears to be an older, incomplete placeholder function for background 
#          processing before `run_upload_analysis_job` was fully built out.
# -----------------------------------------------------------------------------
def run_analysis_job(job_id: str, username: str, doc_files, all_files):
    try:
        # mark running
        JOBS[job_id] = {"status": "running", "message": "", "username": username}

        # ---- DO YOUR EXISTING ANALYSIS LOGIC HERE ----
        # IMPORTANT: reuse your existing code but WITHOUT flash/redirect/render
        # You can call analyze_single_audio_for_ui(...) / analyze_single_text_for_ui(...)
        # and create_notification(...)

        # After finished:
        JOBS[job_id] = {
            "status": "done",
            "message": "Sentiment result is ready ✅",
            "username": username
        }

    except Exception as e:
        JOBS[job_id] = {
            "status": "error",
            "message": str(e),
            "username": username
        }

# -----------------------------------------------------------------------------
# FUNCTION: build_dashboard_data
# PURPOSE: The "Brain" of the Dashboard. It aggregates the chart data (lines, 
#          donuts, bars) and calculates percentages so the HTML template just 
#          has to paint the visuals.
# -----------------------------------------------------------------------------
def build_dashboard_data(username, period, source_type):
    now = datetime.now()

    if period:
        try:
            selected_year, selected_month = map(int, period.split("-"))
        except ValueError:
            selected_year, selected_month = now.year, now.month
    else:
        selected_year, selected_month = now.year, now.month

    MONTH_NAMES = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    month_labels = MONTH_NAMES[:]

    # Line chart (yearly)
    line_complaint, line_non = get_yearly_sentiment_overview(username, selected_year, source_type)

    # Donut (selected month)
    month_c = line_complaint[selected_month - 1]
    month_n = line_non[selected_month - 1]
    total = month_c + month_n
    pct_c = round((month_c / total) * 100) if total else 0
    pct_n = round((month_n / total) * 100) if total else 0

    # Scenario overview (selected month)
    scenario_labels, scenario_c, scenario_n = get_scenario_overview(
        username=username,
        selected_year=selected_year,
        selected_month=selected_month,
        source_type=source_type
    )

    return {
        "month_labels": month_labels,
        "line_complaint": line_complaint,
        "line_non": line_non,
        "pct_complaint": pct_c,
        "pct_non": pct_n,
        "scenario_labels": scenario_labels,
        "scenario_complaint": scenario_c,
        "scenario_non": scenario_n,
        "period": f"{selected_year}-{selected_month:02d}",
        "source_type": source_type,
    }


# ========================
# AUTH ROUTES
# ========================

# -----------------------------------------------------------------------------
# FUNCTION: login
# ROUTE: /login
# PURPOSE: The webpage where users enter their credentials. It checks the DB, 
#          establishes their secure session, and redirects to the right dashboard.
# -----------------------------------------------------------------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        # ✅ which button was clicked ("ADMIN" or "USER")
        selected_role = request.form.get("login_role", "").strip().upper()

        # Optional safety: if someone submits without clicking a button
        if selected_role not in ("ADMIN", "USER"):
            flash("Please choose Admin or User login.")
            return render_template("login.html")

        # 1) Try DB login first (admin-created accounts)
        db_user = fetch_user_by_username(username)
        if db_user and db_user.get("password_hash"):
            if check_password_hash(db_user["password_hash"], password):

                # ✅ block if role doesn't match the button
                if db_user.get("role", "").upper() != selected_role:
                    flash("Access denied for this login button.")
                    return render_template("login.html")

                session["logged_in"] = True
                session["username"] = db_user["username"]
                session["role"] = db_user["role"]

                flash("Logged in successfully.")
                if session["role"] == "ADMIN":
                    return redirect(url_for("admin_home"))
                return redirect(url_for("home"))

        # 2) Fallback to hardcoded USERS
        user = USERS.get(username)
        if user and user["password"] == password:

            # ✅ block if role doesn't match the button
            if user.get("role", "").upper() != selected_role:
                flash("Access denied for this login button.")
                return render_template("login.html")

            session["logged_in"] = True
            session["username"] = username
            session["role"] = user["role"]

            upsert_user_account(username, user["role"], user["password"])
            flash("Logged in successfully.")

            if user["role"] == "ADMIN":
                return redirect(url_for("admin_home"))
            return redirect(url_for("home"))

        flash("Invalid username or password.")
        return render_template("login.html")

    return render_template("login.html")

# -----------------------------------------------------------------------------
# FUNCTION: admin_home
# ROUTE: /admin
# PURPOSE: A simple redirect for admins so they land on their specific dashboard.
# -----------------------------------------------------------------------------
@app.route("/admin")
@admin_required
def admin_home():
    return redirect(url_for("admin_dashboard"))

# -----------------------------------------------------------------------------
# FUNCTION: admin_account_list
# ROUTE: /admin/accounts
# PURPOSE: Renders the HTML page where admins can see a table of all staff accounts.
# -----------------------------------------------------------------------------
@app.route("/admin/accounts")
@admin_required
def admin_account_list():
    return render_template("admin/account_list.html")

# -----------------------------------------------------------------------------
# FUNCTION: admin_account_edit
# ROUTE: /admin/accounts/edit/<int:user_id>
# PURPOSE: Renders the HTML form specifically used for editing an existing staff member.
# -----------------------------------------------------------------------------
@app.route("/admin/accounts/edit/<int:user_id>")
@admin_required
def admin_account_edit(user_id):
    # fetch one user row
    row = fetch_user_account_by_id(user_id)
    if not row:
        abort(404)
    return render_template("admin/account_edit.html", row=row)

# -----------------------------------------------------------------------------
# FUNCTION: api_admin_account
# ROUTE: /api/admin/accounts (GET)
# PURPOSE: An API endpoint that feeds raw user data (JSON) to the Admin Accounts table.
# -----------------------------------------------------------------------------
@app.get("/api/admin/accounts")
@admin_required
def api_admin_account():
    q = (request.args.get("q", "") or "").strip()
    rows = fetch_all_user_account(q)
    return jsonify({"ok": True, "rows": rows})

# -----------------------------------------------------------------------------
# FUNCTION: api_admin_create_account
# ROUTE: /api/admin/accounts (POST)
# PURPOSE: API endpoint that receives data from the "Create User" form and tells 
#          the DB to create the account securely.
# -----------------------------------------------------------------------------
@app.post("/api/admin/accounts")
@admin_required
def api_admin_create_account():
    data = request.get_json(silent=True) or {}

    username = (data.get("username") or "").strip()
    full_name = (data.get("full_name") or "").strip()
    email = (data.get("email") or "").strip().lower()
    password = data.get("password") or ""
    role = (data.get("role") or "USER").strip().upper()

    if role not in ("ADMIN", "USER"):
        role = "USER"

    if not username or not full_name or not email or not password:
        return jsonify({"ok": False, "error": "All fields are required."}), 400

    if len(password) < 8:
        return jsonify({"ok": False, "error": "Password must be at least 8 characters."}), 400

    u_exists, e_exists = username_or_email_exists(username, email)
    if u_exists:
        return jsonify({"ok": False, "error": "Username already exists."}), 409
    if e_exists:
        return jsonify({"ok": False, "error": "Email already exists."}), 409

    new_id = create_user_account(username, full_name, email, role, password)
    return jsonify({"ok": True, "id": new_id})

# -----------------------------------------------------------------------------
# FUNCTION: api_admin_update_account
# ROUTE: /api/admin/accounts/<int:user_id> (PUT)
# PURPOSE: API endpoint to receive changes made to a user and save them via DB.
# -----------------------------------------------------------------------------
@app.put("/api/admin/accounts/<int:user_id>")
@admin_required
def api_admin_update_account(user_id):
    data = request.get_json(silent=True) or {}

    full_name = (data.get("full_name") or "").strip()
    email = (data.get("email") or "").strip()
    role = (data.get("role") or "").strip().upper()

    if role not in ("ADMIN", "USER"):
        return jsonify({"ok": False, "error": "Invalid role"}), 400

    update_user_account(user_id, full_name, email, role)
    return jsonify({"ok": True})

# -----------------------------------------------------------------------------
# FUNCTION: api_admin_delete_account
# ROUTE: /api/admin/accounts/<int:user_id> (DELETE)
# PURPOSE: API endpoint to receive delete requests from the frontend and remove the user.
# -----------------------------------------------------------------------------
@app.delete("/api/admin/accounts/<int:user_id>")
@admin_required
def api_admin_delete_account(user_id):
    try:
        delete_user_account(user_id)
        return jsonify({"ok": True})
    except Exception as e:
        print(f"[DB ERROR] Failed to delete user {user_id}:", e)
        return jsonify({"ok": False, "error": str(e)}), 500

# -----------------------------------------------------------------------------
# FUNCTION: logout
# ROUTE: /logout
# PURPOSE: Destroys the user's secure session cookie, logging them out of the app.
# -----------------------------------------------------------------------------
@app.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out.")
    return redirect(url_for("login"))


# ========================
# MAIN APP ROUTES
# ========================

# -----------------------------------------------------------------------------
# FUNCTION: download_excel
# ROUTE: /download_excel
# PURPOSE: Triggers a direct download of the master `Scenario EIMY.xlsx` file.
# -----------------------------------------------------------------------------
@app.route("/download_excel")
@login_required
def download_excel():
    try:
        return send_file(EXCEL_PATH, as_attachment=True)
    except FileNotFoundError:
        flash("Excel file not found on server.")
        return redirect(url_for("home"))

# -----------------------------------------------------------------------------
# FUNCTION: home
# ROUTE: /
# PURPOSE: Renders the main homepage (which appears to be the Upload page).
# -----------------------------------------------------------------------------
@app.route("/", methods=["GET"])
@login_required
def home():
    # 1) Get all database rows for the logged-in user
    rows = load_history_rows()
    
    # ✅ FLIP THE LIST! Now the newest uploads are at the very top.
    rows.reverse()  
    
    # 2) Grab just the 5 most recent uploads
    recent_rows = rows[:5]
    
    # 3) Format the dates so they look nice on the frontend
    results = []
    for r in recent_rows:
        dt = r.get("datetime")
        dt_str = dt.strftime("%d %b %Y, %I:%M %p") if isinstance(dt, datetime) else str(dt or "")
        
        results.append({
            "filename": r.get("audio_file") or "Unknown",
            "datetime": dt_str,
            "status": "Success"
        })

    # 4) Send the SQL data to the Upload page
    return render_template("upload.html", results=results)

# -----------------------------------------------------------------------------
# FUNCTION: dashboard
# ROUTE: /dashboard
# PURPOSE: Assembles data for a specific user and renders the standard User 
#          Dashboard with charts and graphs.
# -----------------------------------------------------------------------------
@app.route("/dashboard")
@login_required
def dashboard():
    period = request.args.get("period", "")               
    source_type = request.args.get("source_type", "")
    username = session.get("username") 

    now = datetime.now()
    if period:
        try:
            selected_year, selected_month = map(int, period.split("-"))
        except ValueError:
            selected_year, selected_month = now.year, now.month
    else:
        selected_year, selected_month = now.year, now.month

    MONTH_NAMES = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    month_labels = MONTH_NAMES[:]

    username = session.get("username")

    # ✅ Line chart from SQL Server (user-specific)
    line_complaint, line_non = get_yearly_sentiment_overview(username, selected_year, source_type)

    # ✅ Donut from selected month
    month_c = line_complaint[selected_month - 1]
    month_n = line_non[selected_month - 1]
    total_all = month_c + month_n
    pct_c = round((month_c / total_all) * 100) if total_all else 0
    pct_n = round((month_n / total_all) * 100) if total_all else 0

    # ✅ Scenario overview already from SQL Server (user-specific)
    scenario_labels, scenario_c, scenario_n = get_scenario_overview(
        username=username,
        selected_year=selected_year,
        selected_month=selected_month,
        source_type=source_type
    )

    dashboard_data = {
        "username": username,
        "month_labels": month_labels,
        "line_complaint": line_complaint,
        "line_non": line_non,
        "pct_complaint": pct_c,
        "pct_non": pct_n,
        "scenario_labels": scenario_labels,
        "scenario_complaint": scenario_c,
        "scenario_non": scenario_n,
        "period": f"{selected_year}-{selected_month:02d}",
        "source_type": source_type,
    }

    return render_template(
        "dashboard.html",
        dashboard_data=dashboard_data,
        dashboard_action=url_for("dashboard"),
        is_admin=False
    )

# -----------------------------------------------------------------------------
# FUNCTION: admin_dashboard
# ROUTE: /admin/dashboard
# PURPOSE: A special dashboard for Admins. It lets them pick any user from a 
#          dropdown and view that specific user's charts and stats.
# -----------------------------------------------------------------------------
@app.route("/admin/dashboard")
@admin_required
def admin_dashboard():
    period = request.args.get("period", "")
    source_type = request.args.get("source_type", "")
    view_user = request.args.get("username", "")

    users_list = [u for u, info in USERS.items() if info.get("role") != "ADMIN"]
    if not view_user and users_list:
        view_user = users_list[0]

    now = datetime.now()
    if period:
        try:
            selected_year, selected_month = map(int, period.split("-"))
        except ValueError:
            selected_year, selected_month = now.year, now.month
    else:
        selected_year, selected_month = now.year, now.month

    MONTH_NAMES = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    month_labels = MONTH_NAMES[:]

    line_complaint, line_non = get_yearly_sentiment_overview(view_user, selected_year, source_type)

    month_c = line_complaint[selected_month - 1]
    month_n = line_non[selected_month - 1]
    total = month_c + month_n
    pct_c = round((month_c / total) * 100) if total else 0
    pct_n = round((month_n / total) * 100) if total else 0

    scenario_labels, scenario_c, scenario_n = get_scenario_overview(
        username=view_user,
        selected_year=selected_year,
        selected_month=selected_month,
        source_type=source_type
    )

    dashboard_data = {
        "view_user": view_user, 
        "month_labels": month_labels,
        "line_complaint": line_complaint,
        "line_non": line_non,
        "pct_complaint": pct_c,
        "pct_non": pct_n,
        "scenario_labels": scenario_labels,
        "scenario_complaint": scenario_c,
        "scenario_non": scenario_n,
        "period": f"{selected_year}-{selected_month:02d}",
        "source_type": source_type,
        "view_user": view_user,
    }

    return render_template(
        "admin/dashboard.html",
        dashboard_data=dashboard_data,
        dashboard_action=url_for("admin_dashboard"),
        is_admin=True,
        users_list=users_list
    )

# -----------------------------------------------------------------------------
# FUNCTION: comment_page
# ROUTE: /comment/<int:row_id>
# PURPOSE: Renders a form where users can attach custom text comments to a past 
#          AI result. Also handles saving that comment back to Excel/DB.
# -----------------------------------------------------------------------------
@app.route("/comment/<int:row_id>", methods=["GET", "POST"])
@login_required
def comment_page(row_id):
    row = get_history_entry(row_id)
    if not row:
        abort(404)

    if request.method == "POST":
        comment = request.form.get("comment", "").strip()
        update_history_entry(row_id, {"comment": comment})

        # ✅ If request is AJAX, return JSON (no redirect)
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"success": True})

        # ✅ Normal form submit fallback
        flash("Comment saved.")
        return redirect(url_for("sentiment_result"))

    return render_template("comment.html", row=row, row_id=row_id)

# -----------------------------------------------------------------------------
# FUNCTION: upload_file
# ROUTE: /upload
# PURPOSE: The old, synchronous file uploader. It takes files directly from the UI, 
#          sends them to the AI, and waits (which could freeze the page for big files) 
#          before showing the results. (Often replaced by `/upload_async`).
# -----------------------------------------------------------------------------
@app.route("/upload", methods=["POST"])
@login_required
def upload_file():
    results = []
    any_error = False

    # 0) Document uploads (.pdf / .docx)
    doc_files = request.files.getlist("doc_files")
    doc_files = [f for f in doc_files if f and f.filename]

    # 1) Files from multi-select
    files_from_multi = request.files.getlist("audio_files")
    files_from_multi = [f for f in files_from_multi if f and f.filename]

    # 2) Files from folder upload (webkitdirectory)
    files_from_folder = request.files.getlist("audio_folder")
    files_from_folder = [f for f in files_from_folder if f and f.filename]

    # 3) Combine all audio sources
    all_files = files_from_multi + files_from_folder

    # ===== Process PDF/DOCX files =====
    for f in doc_files:
        filename = f.filename

        if not allowed_doc_file(filename):
            flash(f"Skipping '{filename}': only .pdf and .docx are allowed.")
            any_error = True
            continue

        ext = filename.rsplit(".", 1)[1].lower()

        # Extract text
        try:
            if ext == "pdf":
                text = extract_text_from_pdf(f)
            else:
                text = extract_text_from_docx(f)
        except Exception as e:
            flash(f"Error reading document '{filename}': {e}")
            any_error = True
            continue

        if not text.strip():
            flash(f"No readable text found in '{filename}'. If it is a scanned PDF, OCR is required.")
            any_error = True
            continue

        # ✅ Analyze document with Gemini + Scenario
        data = analyze_single_text_for_ui(text, os.path.basename(filename))

        if not data.get("success"):
            flash(f"Error analyzing document '{filename}': {data.get('error', 'Unknown error')}")
            any_error = True
            continue

        create_notification(
           username=session.get("username"),
           message=f"Document analyzed: {os.path.basename(filename)} ({data.get('sentiment')})",
           url="/sentiment_result"
        )

        details = (
            f"Tone: {data.get('tone')} | "
            f"Scenario: ID {data.get('scenario_id')} - {data.get('scenario_title')}\n"
            f"Reason: {data.get('explanation')}"
        )

        saved_ok = save_sentiment_to_db(
            username=session.get("username"),
            source_type="text",
            filename=os.path.basename(filename),
            text_input=text,
            sentiment=data.get("sentiment"),
            confidence=data.get("score"),
            tone=data.get("tone"),
            scenario_id=data.get("scenario_id"),
            scenario_title=data.get("scenario_title"),
            explanation=data.get("explanation"),
            transcript=data.get("transcript"),
        )
        if not saved_ok:
            flash(f"Document analyzed, but failed to save into SQL database: {os.path.basename(filename)}")
            any_error = True
            continue

        results.append({
            "filename": os.path.basename(filename),
            "label": data.get("sentiment"),
            "score": data.get("score"),
            "details": details,
            "tone": data.get("tone"),
            "scenario_id": data.get("scenario_id"),
            "scenario_title": data.get("scenario_title"),
            "transcript": data.get("transcript"),
        })

    # Nothing uploaded?
    if not all_files and not doc_files:
        flash("Please select at least one .wav file, a folder of .wav files, or a .pdf/.docx document.")
        return redirect(url_for("home"))

    # ===== Process WAV files =====
    for file in all_files:
        filename = file.filename

        if not allowed_file(filename):
            flash(f"Skipping '{filename}': only .wav files are allowed.")
            any_error = True
            continue

        file_path = os.path.join(app.config["UPLOAD_FOLDER"], os.path.basename(filename))
        file.save(file_path)

        try:
            label, score, details, raw_data = analyze_sentiment_from_wav(file_path)

            # (save audio result into SQL Server)
            saved_ok = save_sentiment_to_db(
                username=session.get("username"),
                source_type="audio",   # IMPORTANT: audio, not text
                filename=os.path.basename(filename),
                text_input=(raw_data.get("transcript") or f"[AUDIO] {os.path.basename(filename)}"),
                sentiment=label,
                confidence=score,
                tone=raw_data.get("tone"),
                scenario_id=raw_data.get("scenario_id"),
                scenario_title=raw_data.get("scenario_title"),
                explanation=raw_data.get("explanation"),
                transcript=raw_data.get("transcript"),
            )

            if not saved_ok: 
                flash(f"Audio analyzed, but failed to save into SQL database: {os.path.basename(filename)}")
                any_error = True
                continue

            create_notification(
              username=session.get("username"),
              message=f"Audio analyzed: {os.path.basename(filename)} ({label})",
              url="/sentiment_result"
            )

        except Exception as e:
            flash(f"Error while analyzing '{filename}': {e}")
            any_error = True
            continue
        # ✅ Save transcript into a .txt file (NOT Excel)
        transcript_text = (raw_data.get("transcript") or "").strip()

        safe_name = re.sub(r'[<>:"/\\|?*]', "_", os.path.basename(filename))
        safe_base = os.path.splitext(safe_name)[0]
        txt_path = os.path.join(TRANSCRIPT_FOLDER, f"{safe_base}.txt")

        if transcript_text:
           with open(txt_path, "w", encoding="utf-8") as f:
              f.write(transcript_text)

        results.append({
            "filename": os.path.basename(filename),
            "label": label,
            "score": score,
            "details": details,
            "tone": raw_data.get("tone"),
            "scenario_id": raw_data.get("scenario_id"),
            "scenario_title": raw_data.get("scenario_title"),
            "transcript": raw_data.get("transcript"),
        })

    if not results:
        if not any_error:
            flash("No valid files were processed.")
        return redirect(url_for("home"))

    return render_template("upload.html", results=results)

# ========================
# SIDEBAR PAGES
# ========================

# -----------------------------------------------------------------------------
# FUNCTION: sentiment_result
# ROUTE: /sentiment_result
# PURPOSE: Renders the massive History/Results table page. Includes complex logic 
#          for pagination (e.g. 10 rows per page) and filtering by date/type.
# -----------------------------------------------------------------------------
@app.route("/sentiment_result")
@login_required
def sentiment_result():
    file_type = request.args.get("file_type", "")      # wav/pdf/excel/docx/"" all
    sentiment = request.args.get("sentiment", "")      # Complaint/Non-Complaint/"" all
    start = parse_date(request.args.get("start_date", ""))
    end = parse_date(request.args.get("end_date", ""))
    q = (request.args.get("q", "") or "").strip().lower()

    page = int(request.args.get("page", 1))
    per_page = 10

    rows = load_history_rows()
    rows.reverse()  # newest first
    for i, r in enumerate(rows):
        r["row_id"] = i


    filtered = []
    for r in rows:
        if file_type and r.get("file_type") != file_type:
            continue

        if sentiment and str(r.get("sentiment", "")).strip().lower() != sentiment.strip().lower():
            continue

        dt = r.get("datetime")
        if (start or end):
            if not isinstance(dt, datetime):
                continue
            d = dt.date()
            if start and d < start:
                continue
            if end and d > end:
                continue

        if q and q not in str(r.get("audio_file", "")).lower():
            continue

        filtered.append(r)

    total = len(filtered)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))

    start_i = (page - 1) * per_page
    end_i = start_i + per_page
    page_rows = filtered[start_i:end_i]

    for r in page_rows:
      d, t = format_dt_parts(r.get("datetime"))
      r["date_display"] = d
      r["time_display"] = t

    return render_template(
        "sentiment_result.html",
        rows=page_rows,
        total=total,
        page=page,
        total_pages=total_pages,
        file_type=file_type,
        sentiment=sentiment,
        start_date=request.args.get("start_date", ""),
        end_date=request.args.get("end_date", ""),
        q=request.args.get("q", ""),
    )

# -----------------------------------------------------------------------------
# FUNCTION: audio_stream
# ROUTE: /audio/<int:row_id>
# PURPOSE: Streams an audio file directly to the browser's audio player so a 
#          user can listen to the raw call recording without downloading it.
# -----------------------------------------------------------------------------
@app.route("/audio/<int:row_id>")
@login_required
def audio_stream(row_id):
    # 1) Get the same rows order as the Sentiment Result page
    rows = load_history_rows()
    rows.reverse()  # newest first (matches sentiment_result)

    # 2) Validate row_id
    if row_id < 0 or row_id >= len(rows):
        abort(404)

    row = rows[row_id]

    # 3) Build the audio path from your upload folder + filename
    audio_file = row.get("audio_file") or ""
    if not audio_file:
        abort(404)

    audio_path = os.path.join(app.config["UPLOAD_FOLDER"], os.path.basename(audio_file))

    # 4) Check file exists
    if not os.path.exists(audio_path):
        abort(404)

    # 5) Stream audio
    return send_file(audio_path, mimetype="audio/wav", as_attachment=False)

# -----------------------------------------------------------------------------
# FUNCTION: profile
# ROUTE: /profile
# PURPOSE: Renders the user's Profile page, showing their name, role, and a mini 
#          history table of only their specific past uploads/results.
# -----------------------------------------------------------------------------
@app.route("/profile")
@login_required
def profile():
    username = session.get("username")

    profile_info = USER_PROFILES.get(username, {
        "name": username,
        "position": "Unknown",
        "status": "Unknown",
        "department": "",
        "email": "",
    })

    # Optional filters (same style as Sentiment Result)
    file_type = request.args.get("file_type", "")
    sentiment = request.args.get("sentiment", "")
    start = parse_date(request.args.get("start_date", ""))
    end = parse_date(request.args.get("end_date", ""))
    q = (request.args.get("q", "") or "").strip().lower()

    page = int(request.args.get("page", 1))
    per_page = 10

    rows = load_history_rows()
    rows.reverse()  # newest first
    for i, r in enumerate(rows):
        r["row_id"] = i  # needed for links

    # Filter rows
    filtered = []
    for r in rows:
        if file_type and r.get("file_type") != file_type:
            continue

        if sentiment and str(r.get("sentiment", "")).strip().lower() != sentiment.strip().lower():
            continue

        dt = r.get("datetime")
        if (start or end):
            if not isinstance(dt, datetime):
                continue
            d = dt.date()
            if start and d < start:
                continue
            if end and d > end:
                continue

        if q and q not in str(r.get("audio_file", "")).lower():
            continue

        filtered.append(r)

    total = len(filtered)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    start_i = (page - 1) * per_page
    end_i = start_i + per_page
    page_rows = filtered[start_i:end_i]

    # Build table rows for template
    table_rows = []
    for r in page_rows:
        table_rows.append({
            "row_id": r["row_id"],
            "file_type": (r.get("file_type") or "").upper(),
            "file_name": r.get("audio_file", ""),
            "sentiment": r.get("sentiment", ""),
            "score": r.get("score_display", ""),
            "date_created": format_dt(r.get("datetime")),
            "comment": r.get("comment", ""),  # ✅ edited comment appears here
        })

    return render_template(
        "profile.html",
        profile_info=profile_info,
        rows=table_rows,
        total=total,
        page=page,
        total_pages=total_pages,
        file_type=file_type,
        sentiment=sentiment,
        start_date=request.args.get("start_date", ""),
        end_date=request.args.get("end_date", ""),
        q=request.args.get("q", ""),
    )

# -----------------------------------------------------------------------------
# FUNCTION: change_password
# ROUTE: /change-password
# PURPOSE: Validates and updates a user's password securely within the DB.
# -----------------------------------------------------------------------------
@app.route("/change-password", methods=["POST"])
@login_required
def change_password():
    username = session.get("username")

    current_password = request.form.get("current_password", "").strip()
    new_password = request.form.get("new_password", "").strip()
    confirm_password = request.form.get("confirm_password", "").strip()

    if not current_password or not new_password or not confirm_password:
        flash("All password fields are required.")
        return redirect(url_for("profile"))

    if new_password != confirm_password:
        flash("New password and confirmation do not match.")
        return redirect(url_for("profile"))

    if len(new_password) < 8:
        flash("New password must be at least 8 characters.")
        return redirect(url_for("profile"))

    # 🔐 fetch user from DB
    user = fetch_user_by_username(username)
    if not user or not user.get("password_hash"):
        flash("Account not found.")
        return redirect(url_for("profile"))

    # 🔐 verify current password
    if not check_password_hash(user["password_hash"], current_password):
        flash("Current password is incorrect.")
        return redirect(url_for("profile"))

    # 🔐 update password hash
    new_hash = generate_password_hash(new_password)

    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute(
       """
       UPDATE dbo.user_account
       SET user_password = ?
       WHERE username = ?
       """,
       (new_hash, username)
    )

        conn.commit()
    finally:
        cur.close()
        conn.close()

    flash("Password updated successfully.")
    return redirect(url_for("profile"))

# -----------------------------------------------------------------------------
# FUNCTION: edit_client
# ROUTE: /profile/edit/<file_id>
# PURPOSE: Allows a user to edit the text summary of a past analysis row.
# -----------------------------------------------------------------------------
@app.route("/profile/edit/<file_id>", methods=["GET", "POST"])
@login_required
def edit_client(file_id):
    if request.method == "POST":
        new_summary = request.form.get("summary", "")
        update_history_entry(file_id, {"summary": new_summary})
        return redirect(url_for("profile"))

    row = get_history_entry(file_id)
    return render_template("edit_client.html", row=row)

# -----------------------------------------------------------------------------
# FUNCTION: delete_client
# ROUTE: /profile/delete/<file_id>
# PURPOSE: Endpoint that permanently deletes a past analysis from the system.
# -----------------------------------------------------------------------------
@app.route("/profile/delete/<file_id>", methods=["POST"])
@login_required
def delete_client(file_id):
    delete_history_entry(file_id)
    return redirect(url_for("profile", deleted=1))

# -----------------------------------------------------------------------------
# FUNCTION: delete_sentiment_result
# ROUTE: /sentiment_result/delete/<int:row_id>
# PURPOSE: Endpoint that permanently deletes a past analysis from the Sentiment Result page.
# -----------------------------------------------------------------------------
@app.route("/sentiment_result/delete/<int:row_id>", methods=["POST"])
@login_required
def delete_sentiment_result(row_id):
    delete_history_entry(row_id)
    return redirect(url_for("sentiment_result"))

# -----------------------------------------------------------------------------
# FUNCTION: history
# ROUTE: /history
# PURPOSE: Renders a simplified history view, likely an older version of the UI.
# -----------------------------------------------------------------------------
@app.route("/history")
@login_required
def history():
    rows = load_history_rows()
    rows.reverse()   # Newest first

    history_rows = []
    for idx, row in enumerate(rows):
        history_rows.append({
            "idx": idx,
            "date_time": format_dt(row.get("datetime")),
            "audio_file": row.get("audio_file") or "",
        })

    return render_template("history.html", history=history_rows)

# -----------------------------------------------------------------------------
# FUNCTION: history_pdf
# ROUTE: /history_pdf/<int:row_id>
# PURPOSE: Calls `build_history_pdf` to dynamically create a PDF, then opens it 
#          directly in the user's web browser tab (View Mode).
# -----------------------------------------------------------------------------
@app.route("/history_pdf/<int:row_id>")
@login_required
def history_pdf(row_id):
    rows = load_history_rows()
    rows.reverse()  # 👈 add this so index matches the table

    if row_id < 0 or row_id >= len(rows):
        abort(404)

    pdf_path = build_history_pdf(rows[row_id])
    return send_file(pdf_path, mimetype="application/pdf")

# -----------------------------------------------------------------------------
# FUNCTION: history_pdf_download
# ROUTE: /history_pdf_download/<int:row_id>
# PURPOSE: Generates the PDF, but forces the browser to download it as a file 
#          rather than opening it in a tab.
# -----------------------------------------------------------------------------
@app.route("/history_pdf_download/<int:row_id>")
@login_required
def history_pdf_download(row_id):
    rows = load_history_rows()
    rows.reverse()  # 👈 add this too

    if row_id < 0 or row_id >= len(rows):
        abort(404)

    row = rows[row_id]
    pdf_path = build_history_pdf(row)
    download_name = f"sentiment_{row.get('audio_file', 'history')}.pdf"

    return send_file(
        pdf_path,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=download_name,
    )

# -----------------------------------------------------------------------------
# FUNCTION: result_pdf
# ROUTE: /result_pdf
# PURPOSE: Similar to history_pdf, but locates the record by filename instead 
#          of its index number.
# -----------------------------------------------------------------------------
@app.route("/result_pdf")
@login_required
def result_pdf():
    filename = (request.args.get("filename") or "").strip()
    if not filename:
        abort(400)

    rows = load_history_rows()
    rows.reverse()  # newest first (same as your UI)

    # Find newest match by filename
    target = None
    for r in rows:
        if (r.get("audio_file") or "").strip().lower() == filename.lower():
            target = r
            break

    if not target:
        abort(404)

    pdf_path = build_history_pdf(target)

    # Opens in browser (view mode)
    return send_file(pdf_path, mimetype="application/pdf", as_attachment=False)

# -----------------------------------------------------------------------------
# FUNCTION: download_filtered_excel
# ROUTE: /sentiment_result/download_excel
# PURPOSE: "Export to Excel" button logic. Takes whatever filters the user has 
#          applied to the history table and turns just those rows into an .xlsx file.
# -----------------------------------------------------------------------------
@app.route("/sentiment_result/download_excel")
@login_required
def download_filtered_excel():
    
    from openpyxl import Workbook

    file_type = request.args.get("file_type", "")
    sentiment = request.args.get("sentiment", "")
    start = parse_date(request.args.get("start_date", ""))
    end = parse_date(request.args.get("end_date", ""))
    q = (request.args.get("q", "") or "").strip().lower()

    rows = load_history_rows()
    rows.reverse()

    filtered = []
    for r in rows:
        if file_type and r.get("file_type") != file_type:
            continue
        if sentiment and str(r.get("sentiment", "")).strip().lower() != sentiment.strip().lower():
            continue

        dt = r.get("datetime")
        if (start or end):
            if not isinstance(dt, datetime):
                continue
            d = dt.date()
            if start and d < start:
                continue
            if end and d > end:
                continue

        if q and q not in str(r.get("audio_file", "")).lower():
            continue

        filtered.append(r)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sentiment Result"

    ws.append(["File Type", "File Name", "Summary", "Sentiment", "Sentiment Score", "Date Created"])
    for r in filtered:
        dt = r.get("datetime")
        ws.append([
            r.get("file_type", ""),
            r.get("audio_file", ""),
            r.get("summary", ""),
            r.get("sentiment", ""),
            r.get("score_display", ""),
            dt.strftime("%d %b %Y, %I:%M %p") if isinstance(dt, datetime) else ""
        ])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name="sentiment_result_filtered.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# -----------------------------------------------------------------------------
# FUNCTION: download_filtered_pdf
# ROUTE: /sentiment_result/download_pdf
# PURPOSE: "Export to PDF" button logic. Same idea as the excel export, but 
#          generates a multi-page PDF table of the currently filtered results.
# -----------------------------------------------------------------------------
@app.route("/sentiment_result/download_pdf")
@login_required
def download_filtered_pdf():
    file_type = request.args.get("file_type", "")
    sentiment = request.args.get("sentiment", "")
    start = parse_date(request.args.get("start_date", ""))
    end = parse_date(request.args.get("end_date", ""))
    q = (request.args.get("q", "") or "").strip().lower()

    rows = load_history_rows()
    rows.reverse()

    filtered = []
    for r in rows:
        if file_type and r.get("file_type") != file_type:
            continue
        if sentiment and str(r.get("sentiment", "")).strip().lower() != sentiment.strip().lower():
            continue

        dt = r.get("datetime")
        if (start or end):
            if not isinstance(dt, datetime):
                continue
            d = dt.date()
            if start and d < start:
                continue
            if end and d > end:
                continue

        if q and q not in str(r.get("audio_file", "")).lower():
            continue

        filtered.append(r)

    # Build PDF using reportlab (simple table)
    pdf_path = os.path.join(os.getcwd(), "sentiment_result_filtered.pdf")
    doc = SimpleDocTemplate(pdf_path, pagesize=landscape(A4))

    data = [["File Type", "File Name", "Summary", "Sentiment", "Score", "Date Created"]]
    for r in filtered:
        dt = r.get("datetime")
        data.append([
            (r.get("file_type") or "").upper(),
            r.get("audio_file") or "",
            r.get("summary") or "",
            r.get("sentiment") or "",
            r.get("score_display") or "",
            dt.strftime("%d %b %Y, %I:%M %p") if isinstance(dt, datetime) else ""
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    doc.build([table])

    return send_file(
        pdf_path,
        mimetype="application/pdf",
        as_attachment=True,
        download_name="sentiment_result_filtered.pdf"
    )

# -----------------------------------------------------------------------------
# FUNCTION: transcript_view
# ROUTE: /transcript/<int:row_id>
# PURPOSE: Renders the page where a user can read the full dialogue/transcript 
#          of what was said during an audio call.
# -----------------------------------------------------------------------------
@app.route("/transcript/<int:row_id>")
@login_required
def transcript_view(row_id):
    rows = load_history_rows()
    rows.reverse()

    if row_id < 0 or row_id >= len(rows):
        abort(404)

    row = rows[row_id]
    audio_file = row.get("audio_file", "")

    # ✅ derive transcript .txt name from audio file name
    safe_name = re.sub(r'[<>:"/\\|?*]', "_", os.path.basename(audio_file))
    safe_base = os.path.splitext(safe_name)[0]
    txt_path = os.path.join(TRANSCRIPT_FOLDER, f"{safe_base}.txt")

    transcript_text = ""
    if os.path.exists(txt_path):
        with open(txt_path, "r", encoding="utf-8") as f:
            transcript_text = f.read().strip()

      # ✅ Fallback to transcript stored in Excel (works for PDF/DOCX entries)
    if not transcript_text:
        transcript_text = (row.get("transcript") or "").strip()

    if not transcript_text:
        transcript_text = "No transcript available."

    return render_template(
        "transcript_view.html",
        audio_file=audio_file,
        scenario_title=row.get("scenario_title", ""),
        transcript=transcript_text
    )

# -----------------------------------------------------------------------------
# FUNCTION: download_excel_row
# ROUTE: /excel_row/<int:row_id>
# PURPOSE: Generates a tiny Excel file containing just one specific row of data.
# -----------------------------------------------------------------------------
@app.route("/excel_row/<int:row_id>")
@login_required
def download_excel_row(row_id):
    from openpyxl import Workbook
    from io import BytesIO

    rows = load_history_rows()
    rows.reverse()

    if row_id < 0 or row_id >= len(rows):
        abort(404)

    r = rows[row_id]

    wb = Workbook()
    ws = wb.active
    ws.title = "Client Result"

    ws.append([
        "Date/Time", "File Name", "Sentiment", "Score", "Tone",
        "Scenario ID", "Scenario Title", "Explanation", "Transcript", "Comment"
    ])

    dt = r.get("datetime")
    dt_str = dt.strftime("%d %b %Y, %I:%M %p") if isinstance(dt, datetime) else str(dt or "")

    ws.append([
        dt_str,
        r.get("audio_file", ""),
        r.get("sentiment", ""),
        r.get("score_display", ""),
        r.get("tone", ""),
        r.get("scenario_id", ""),
        r.get("scenario_title", ""),
        r.get("explanation", ""),
        r.get("transcript", ""),
        r.get("comment", ""),
    ])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    safe_name = re.sub(r'[<>:"/\\|?*]', "_", r.get("audio_file", "result"))
    safe_name = os.path.splitext(safe_name)[0]

    return send_file(
        bio,
        as_attachment=True,
        download_name=f"{safe_name}_result.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# -----------------------------------------------------------------------------
# FUNCTION: download_selected_excel
# ROUTE: /download_selected_excel
# PURPOSE: Uses checkboxes in the UI! If a user checks 3 specific rows, this 
#          takes those IDs and generates an Excel file of just those 3 rows.
# -----------------------------------------------------------------------------
@app.route("/download_selected_excel", methods=["POST"])
@login_required
def download_selected_excel():
    selected_ids = request.form.getlist("selected_ids")
    if not selected_ids:
        return ("No rows selected", 400)

    # Use the same rows/order as the Sentiment Result page
    rows = load_history_rows()
    rows.reverse()

    try:
        selected_ids = [int(x) for x in selected_ids]
    except ValueError:
        return ("Invalid selection", 400)

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Selected Results"

    # Header
    out_ws.append(["File Type", "File Name", "Summary", "Sentiment", "Sentiment Score", "Date Created"])

    for rid in selected_ids:
        if rid < 0 or rid >= len(rows):
            continue
        r = rows[rid]
        dt = r.get("datetime")
        out_ws.append([
            (r.get("file_type") or "").upper(),
            r.get("audio_file") or "",
            r.get("summary") or "",
            r.get("sentiment") or "",
            r.get("score_display") or "",
            dt.strftime("%d/%m/%Y %I:%M %p") if isinstance(dt, datetime) else ""
        ])

    bio = BytesIO()
    out_wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name="selected_results.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# -----------------------------------------------------------------------------
# FUNCTION: download_selected_pdf
# ROUTE: /download_selected_pdf
# PURPOSE: Uses checkboxes! Select a few rows, and this turns those specific 
#          analyses into a combined multi-page PDF document.
# -----------------------------------------------------------------------------
@app.route("/download_selected_pdf", methods=["POST"])
@login_required
def download_selected_pdf():
    selected_ids = request.form.getlist("selected_ids")
    if not selected_ids:
        return ("No rows selected", 400)

    # IMPORTANT: use the same rows list as your UI
    rows = load_history_rows()
    rows.reverse()  # match Sentiment Result page order (newest first)

    # Convert selected ids to ints safely
    try:
        selected_ids = [int(x) for x in selected_ids]
    except ValueError:
        return ("Invalid selection", 400)

    styles = getSampleStyleSheet()
    body = styles["BodyText"]
    body.wordWrap = "CJK"  # good wrapping

    def P(text):
        return Paragraph(str(text or "").replace("\n", "<br/>"), body)

    bio = BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=landscape(A4),
        leftMargin=30,
        rightMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    story = []
    first = True

    for rid in selected_ids:
        if rid < 0 or rid >= len(rows):
            continue

        row = rows[rid]

        # new page between results
        if not first:
            story.append(PageBreak())
        first = False

        # Title (same style as your single PDF)
        story.append(Paragraph("CLIENT RESULT", styles["Title"]))
        story.append(Spacer(1, 12))

        data = [
            ["Field",          "Value"],
            ["Date / Time",    P(format_dt(row.get("datetime")))],
            ["File Name",      P(row.get("audio_file"))],
            ["Sentiment",      P(row.get("sentiment"))],
            ["Score (%)",      P(row.get("score_display"))],
            ["Tone",           P(row.get("tone"))],
            ["Scenario ID",    P(row.get("scenario_id"))],
            ["Scenario Title", P(row.get("scenario_title"))],
            ["Explanation",    P(row.get("explanation"))],
            ["Transcript",     P(row.get("transcript"))],
            ["Comment / Feedback", P(row.get("comment") or "")],
        ]

        table = Table(data, colWidths=[120, 650], repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ]))

        story.append(table)

    # if nothing valid selected
    if not story:
        return ("No valid rows selected", 400)

    doc.build(story)

    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name="selected_results.pdf",
        mimetype="application/pdf",
    )

# -----------------------------------------------------------------------------
# FUNCTION: api_unread_count
# ROUTE: /api/notifications/unread-count
# PURPOSE: An invisible API called constantly by the webpage in the background 
#          to refresh the notification bell number without reloading the page.
# -----------------------------------------------------------------------------
@app.get("/api/notifications/unread-count")
@login_required
def api_unread_count():
    username = session.get("username")
    return jsonify({"unread_count": unread_count(username)})

# -----------------------------------------------------------------------------
# FUNCTION: api_notifications
# ROUTE: /api/notifications
# PURPOSE: Called when the user clicks the notification bell. It returns the actual 
#          messages (e.g. "Analysis Complete") so the dropdown list populates.
# -----------------------------------------------------------------------------
@app.get("/api/notifications")
@login_required
def api_notifications():
    username = session.get("username")
    limit = int(request.args.get("limit", 10))

    items = load_notifications(username, limit=limit)
    return jsonify({
        "unread_count": unread_count(username),
        "items": items
    })

# -----------------------------------------------------------------------------
# FUNCTION: api_mark_read
# ROUTE: /api/notifications/mark-read
# PURPOSE: Triggered when a user clicks on one specific notification to read it.
# -----------------------------------------------------------------------------
@app.post("/api/notifications/mark-read")
@login_required
def api_mark_read():
    username = session.get("username")
    data = request.get_json(silent=True) or {}
    notif_id = data.get("id")

    if notif_id is None:
        return jsonify({"ok": False, "error": "Missing id"}), 400

    mark_notification_read(username, int(notif_id))
    return jsonify({"ok": True})

# -----------------------------------------------------------------------------
# FUNCTION: api_mark_all_read
# ROUTE: /api/notifications/mark-all-read
# PURPOSE: Triggered by the "Clear All / Mark all as read" button in notifications.
# -----------------------------------------------------------------------------
@app.post("/api/notifications/mark-all-read")
@login_required
def api_mark_all_read():
    username = session.get("username")
    mark_all_notifications_read(username)
    return jsonify({"ok": True})

# -----------------------------------------------------------------------------
# FUNCTION: upload_async
# ROUTE: /upload_async
# PURPOSE: The modern version of the file uploader. It takes the files, immediately 
#          tells the user "Got it!", and starts the AI process on a separate thread 
#          so the user can navigate away while it finishes.
# -----------------------------------------------------------------------------
@app.post("/upload_async")
@login_required
def upload_async():
    username = session.get("username")

    # Collect files the same way as /upload
    doc_files = request.files.getlist("doc_files")
    doc_files = [f for f in doc_files if f and f.filename]

    files_from_multi = request.files.getlist("audio_files")
    files_from_multi = [f for f in files_from_multi if f and f.filename]

    files_from_folder = request.files.getlist("audio_folder")
    files_from_folder = [f for f in files_from_folder if f and f.filename]

    all_files = files_from_multi + files_from_folder

    if not all_files and not doc_files:
        return jsonify({"ok": False, "error": "No files selected"}), 400

    # Save audio files to disk (so thread can read them later)
    audio_paths = []
    for f in all_files:
        if not allowed_file(f.filename):
            continue
        file_path = os.path.join(app.config["UPLOAD_FOLDER"], os.path.basename(f.filename))
        f.save(file_path)
        audio_paths.append(file_path)

    # Store doc files in memory (bytes) for thread processing
    doc_files_meta = []
    for f in doc_files:
        if not allowed_doc_file(f.filename):
            continue
        ext = f.filename.rsplit(".", 1)[1].lower()
        doc_files_meta.append({
            "filename": f.filename,
            "ext": ext,
            "bytes": f.read()
        })

    job_id = str(uuid.uuid4())

    with JOBS_LOCK:
        JOBS[job_id] = {"status": "running", "message": "Processing...", "username": username}

    # Run analysis in background
    t = Thread(
        target=run_upload_analysis_job,
        args=(job_id, username, doc_files_meta, audio_paths),
        daemon=True
    )
    t.start()

    # Save job_id so /api/job_status can work without passing query param too
    session["last_job_id"] = job_id

    return jsonify({"ok": True, "job_id": job_id})

# -----------------------------------------------------------------------------
# FUNCTION: api_job_status
# ROUTE: /api/job_status
# PURPOSE: Used with `upload_async`. The website constantly calls this API 
#          to ask "Is it done yet?" so it can update the progress bar.
# -----------------------------------------------------------------------------
@app.get("/api/job_status")
@login_required
def api_job_status():
    job_id = request.args.get("job_id") or session.get("last_job_id")
    if not job_id:
        return jsonify({"status": "none"})

    with JOBS_LOCK:
        job = JOBS.get(job_id)

    if not job:
        return jsonify({"status": "none"})

    if job.get("username") != session.get("username"):
        return jsonify({"status": "none"})

    return jsonify({"status": job.get("status"), "message": job.get("message", "")})

# -----------------------------------------------------------------------------
# FUNCTION: vapid_public_key
# ROUTE: /vapidPublicKey
# PURPOSE: Required for Web Push Notifications. Hands the browser the secure 
#          public key it needs to subscribe to alerts.
# -----------------------------------------------------------------------------
@app.get("/vapidPublicKey")
@login_required
def vapid_public_key():
    if not VAPID_PUBLIC_KEY:
        return ("VAPID_PUBLIC_KEY not set", 500)
    return Response(VAPID_PUBLIC_KEY, mimetype="text/plain")

# -----------------------------------------------------------------------------
# FUNCTION: save_subscription
# ROUTE: /saveSubscription
# PURPOSE: Receives the user's browser details and saves them via `save_push_subs` 
#          so the server can send them web push notifications later.
# -----------------------------------------------------------------------------
@app.post("/saveSubscription")
@login_required
def save_subscription():
    sub = request.get_json(force=True)
    username = session.get("username")

    with PUSH_LOCK:
        PUSH_SUBSCRIPTIONS[username] = sub
        save_push_subs(PUSH_SUBSCRIPTIONS)

    return jsonify({"ok": True})

# -----------------------------------------------------------------------------
# FUNCTION: debug_subs
# ROUTE: /debug_subs
# PURPOSE: An admin/developer tool to quickly check if push notifications are 
#          correctly registered for the current user.
# -----------------------------------------------------------------------------
@app.get("/debug_subs")
@login_required
def debug_subs():
    username = session.get("username")
    return jsonify({
        "current_user": username,
        "has_subscription": username in PUSH_SUBSCRIPTIONS,
        "saved_users": list(PUSH_SUBSCRIPTIONS.keys())
    })


# ========================
# Run app
# ========================
if __name__ == "__main__":
    app.run(debug=True, use_reloader=False)